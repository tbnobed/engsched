from flask import render_template, redirect, url_for, flash, request, jsonify, send_file, make_response, Response
from flask_login import login_user, logout_user, login_required, current_user
from app import app, db, is_mobile_device, csrf
from models import User, Schedule, QuickLink, Location, EmailSettings, TicketCategory, Ticket, TicketComment, TicketHistory, TicketStatus, RecurringScheduleTemplate, TicketView
from forms import (
    LoginForm, RegistrationForm, ScheduleForm, AdminUserForm, EditUserForm, 
    ChangePasswordForm, QuickLinkForm, LocationForm, EmailSettingsForm, RecurringScheduleForm
)
from datetime import datetime, timedelta, time
import pytz
import csv
import random
import string
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
import json
import os
import requests
from werkzeug.utils import secure_filename
from email_utils import send_schedule_notification
from flask import session

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """New landing page with tickets, studio bookings, and technician schedules"""
    # Check for force mobile parameter for testing
    force_mobile = request.args.get('mobile') == 'true'
    
    # Check if user is on mobile device and redirect to mobile calendar
    if is_mobile_device() or force_mobile:
        app.logger.debug("Mobile device detected in dashboard - redirecting to calendar")
        return redirect(url_for('calendar', mobile='true' if force_mobile else None))
    
    app.logger.debug("Desktop device detected in dashboard - serving desktop template")
    user_tz = pytz.timezone(current_user.get_timezone())
    today = datetime.now(user_tz).date()
    
    # Get today's date range in UTC for database queries
    today_start_local = user_tz.localize(datetime.combine(today, time.min))
    today_end_local = user_tz.localize(datetime.combine(today, time.max))
    today_start_utc = today_start_local.astimezone(pytz.UTC)
    today_end_utc = today_end_local.astimezone(pytz.UTC)
    
    # Get relevant ticket information with unread activity detection
    recent_tickets = Ticket.query.filter(
        Ticket.status.in_(['open', 'in_progress', 'pending'])
    ).order_by(Ticket.priority.desc(), Ticket.created_at.desc()).limit(10).all()
    
    # Add unread activity indicators for dashboard tickets
    for ticket in recent_tickets:
        try:
            # Check if this user has viewed this ticket
            ticket_view = TicketView.query.filter_by(
                user_id=current_user.id,
                ticket_id=ticket.id
            ).first()
            
            if ticket_view:
                # Check for activity since last view
                ticket.has_unread_activity = (
                    ticket.updated_at > ticket_view.last_viewed_at or
                    TicketComment.query.filter(
                        TicketComment.ticket_id == ticket.id,
                        TicketComment.created_at > ticket_view.last_viewed_at
                    ).count() > 0 or
                    TicketHistory.query.filter(
                        TicketHistory.ticket_id == ticket.id,
                        TicketHistory.created_at > ticket_view.last_viewed_at
                    ).count() > 0
                )
            else:
                # User has never viewed this ticket - it's new to them
                ticket.has_unread_activity = True
        except Exception as e:
            app.logger.error(f"Error checking unread activity for ticket {ticket.id}: {e}")
            # Default to showing as unread if there's an error
            ticket.has_unread_activity = True
    
    # Get today's technician schedules using overlap logic
    raw_schedules = Schedule.query.filter(
        Schedule.start_time <= today_end_utc,
        Schedule.end_time >= today_start_utc
    ).order_by(Schedule.start_time).all()
    
    # Apply timezone-aware filtering for all-day OOO entries to prevent date shifting
    today_schedules = []
    for schedule in raw_schedules:
        # Special handling for all-day time-off events to prevent timezone date shifting
        if schedule.time_off and schedule.all_day:
            # For all-day events, we need to determine the intended calendar date
            # Since existing entries were created in Chicago time, we reverse-engineer the date
            utc_time = schedule.start_time.astimezone(pytz.UTC)
            
            # Try to determine the original calendar date by checking common US timezones
            chicago_tz = pytz.timezone('America/Chicago')
            pacific_tz = pytz.timezone('America/Los_Angeles')
            
            chicago_display = utc_time.astimezone(chicago_tz)
            pacific_display = utc_time.astimezone(pacific_tz)
            
            # If the UTC time matches Chicago midnight conversion pattern, use Chicago date
            chicago_midnight = chicago_tz.localize(datetime.combine(chicago_display.date(), time(0, 0)))
            if utc_time == chicago_midnight.astimezone(pytz.UTC):
                intended_date = chicago_display.date()
                app.logger.debug(f"Dashboard all-day OOO {schedule.id}: Detected Chicago-created entry for {intended_date}")
            else:
                # Otherwise, use the current user's timezone date
                intended_date = utc_time.astimezone(user_tz).date()
                app.logger.debug(f"Dashboard all-day OOO {schedule.id}: Using user timezone date {intended_date}")
            
            # Only include if the intended date matches today's date
            if intended_date == today:
                # Apply timezone-aware display time conversion
                schedule.start_time = user_tz.localize(datetime.combine(intended_date, time(0, 0)))
                schedule.end_time = user_tz.localize(datetime.combine(intended_date, time(23, 59)))
                today_schedules.append(schedule)
                app.logger.debug(f"Dashboard all-day OOO {schedule.id}: Included for {intended_date}")
            else:
                app.logger.debug(f"Dashboard all-day OOO {schedule.id}: Filtered out - intended date {intended_date} != today {today}")
        else:
            # Regular schedules - include normally with timezone conversion
            schedule.start_time = schedule.start_time.astimezone(user_tz)
            schedule.end_time = schedule.end_time.astimezone(user_tz)
            today_schedules.append(schedule)
    
    # Organize schedules by technician  
    schedules_by_tech = {}
    for schedule in today_schedules:
        tech_name = schedule.technician.username
        if tech_name not in schedules_by_tech:
            schedules_by_tech[tech_name] = {
                'user': schedule.technician,
                'schedules': []
            }
        schedules_by_tech[tech_name]['schedules'].append(schedule)
    
    return render_template('dashboard.html', 
                         recent_tickets=recent_tickets,
                         schedules_by_tech=schedules_by_tech,
                         today=today,
                         user_timezone=current_user.get_timezone())

@app.route('/api/studio-bookings')
@login_required
def api_studio_bookings():
    """Fetch today's studio bookings from the public API"""
    try:
        user_tz = pytz.timezone(current_user.get_timezone())
        today = datetime.now(user_tz)
        
        # Construct dates like JavaScript: setHours(0,0,0,0) and setHours(23,59,59,999)
        start_of_day = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = today.replace(hour=23, minute=59, second=59, microsecond=999000)  # 999ms = 999000μs
        
        # Convert to UTC and format as ISO strings
        start_of_day_utc = start_of_day.astimezone(pytz.UTC)
        end_of_day_utc = end_of_day.astimezone(pytz.UTC)
        
        start_time = start_of_day_utc.isoformat().replace('+00:00', 'Z')
        end_time = end_of_day_utc.isoformat().replace('+00:00', 'Z')
        
        # Make request to studio booking API - configurable base URL
        studio_api_base = os.environ.get('STUDIO_BOOKING_API_URL', 'https://plex.bookstud.io')
        api_url = f"{studio_api_base}/api/public/bookings?start={start_time}&end={end_time}"
        
        app.logger.debug(f"Studio bookings API request: {api_url}")
        
        response = requests.get(api_url, timeout=10)
        response.raise_for_status()
        
        bookings_data = response.json()
        
        return jsonify({
            'success': True,
            'bookings': bookings_data,
            'date': today.strftime('%Y-%m-%d'),
            'debug': {
                'start_time': start_time,
                'end_time': end_time,
                'timezone': current_user.get_timezone()
            }
        })
        
    except requests.exceptions.RequestException as e:
        app.logger.error(f"Error fetching studio bookings: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Unable to fetch studio bookings',
            'message': str(e)
        }), 500
    except Exception as e:
        app.logger.error(f"Unexpected error in studio bookings API: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Unexpected error',
            'message': str(e)
        }), 500

@app.route('/api/inbound-email', methods=['POST'])
@csrf.exempt
def inbound_email_webhook():
    """
    SendGrid Inbound Parse webhook endpoint
    Receives emails sent to mail1.opscal.io and converts them to tickets
    """
    try:
        app.logger.info("Received inbound email webhook from SendGrid")
        
        # Get email data from SendGrid webhook
        email_data = request.form
        
        # Check if we have raw email data
        raw_email = email_data.get('email', '')
        
        if raw_email:
            # Parse the raw MIME email
            import email
            from email.parser import Parser
            
            app.logger.debug("Processing raw MIME email data")
            msg = email.message_from_string(raw_email)
            
            # Extract components from parsed email
            from_email = msg.get('From', 'unknown@email.com')
            subject = msg.get('Subject', 'No Subject')
            to_email = msg.get('To', '')
            
            # Extract body content
            text_content = ''
            html_content = ''
            
            if msg.is_multipart():
                for part in msg.walk():
                    content_type = part.get_content_type()
                    if content_type == 'text/plain':
                        text_content = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                    elif content_type == 'text/html':
                        html_content = part.get_payload(decode=True).decode('utf-8', errors='ignore')
            else:
                # Single part message
                content_type = msg.get_content_type()
                payload = msg.get_payload(decode=True)
                if payload:
                    content = payload.decode('utf-8', errors='ignore')
                    if content_type == 'text/html':
                        html_content = content
                    else:
                        text_content = content
        else:
            # Fallback to form data parsing (for older SendGrid format)
            from_email = email_data.get('from', 'unknown@email.com')
            subject = email_data.get('subject', 'No Subject')
            text_content = email_data.get('text', '')
            html_content = email_data.get('html', '')
            to_email = email_data.get('to', '')
        
        app.logger.info(f"Email from: {from_email}, subject: {subject}, to: {to_email}")
        app.logger.debug(f"Text content length: {len(text_content)}, HTML content length: {len(html_content)}")
        app.logger.debug(f"Email data fields: {list(email_data.keys())}")
        
        # Log all available fields for debugging
        for key, value in email_data.items():
            if key not in ['text', 'html', 'email']:  # Don't log large content fields
                app.logger.debug(f"Email field '{key}': {str(value)[:200]}...")
        
        # CHECK FOR REPLY TO EXISTING TICKET FIRST
        import re
        ticket_reply_match = re.search(r'\[Ticket #(\d+)\]', subject, re.IGNORECASE)
        if ticket_reply_match:
            ticket_id = int(ticket_reply_match.group(1))
            app.logger.info(f"Detected reply to existing ticket #{ticket_id}")
            
            # Find the existing ticket
            existing_ticket = Ticket.query.get(ticket_id)
            if existing_ticket:
                app.logger.info(f"Found existing ticket #{ticket_id}, adding comment instead of creating new ticket")
                
                # Clean up the description for the comment
                description = text_content.strip() if text_content else ""
                if not description and html_content:
                    import re
                    html_clean = re.sub('<[^<]+?>', '', html_content)
                    description = html_clean.strip()
                
                if not description:
                    description = "Reply received with no readable content"
                
                # Extract sender info
                sender_name = from_email
                if '<' in from_email:
                    sender_name = from_email.split('<')[0].strip()
                    from_email = from_email.split('<')[1].strip('>')
                
                # Try to find the user by email
                sender_user = User.query.filter(User.email.ilike(from_email)).first()
                
                # Add comment to existing ticket
                comment_content = f"Email reply from: {sender_name} ({from_email})\n\n{description}"
                
                if sender_user:
                    # Internal user replying
                    new_comment = TicketComment(
                        ticket_id=existing_ticket.id,
                        user_id=sender_user.id,
                        content=comment_content
                    )
                    commenter_name = sender_user.username
                else:
                    # External user replying - assign comment to admin but note the external sender
                    primary_admin = User.query.filter_by(email='admin@obedtv.com').first()
                    if not primary_admin:
                        primary_admin = User.query.filter_by(is_admin=True).first()
                    
                    new_comment = TicketComment(
                        ticket_id=existing_ticket.id,
                        user_id=primary_admin.id if primary_admin else 1,
                        content=comment_content
                    )
                    commenter_name = f"{sender_name} (external)"
                
                db.session.add(new_comment)
                db.session.commit()
                
                app.logger.info(f"Added comment to ticket #{ticket_id} from {commenter_name}")
                
                # Send comment notification to all relevant parties
                try:
                    from email_utils import send_ticket_comment_notification
                    if sender_user:
                        send_ticket_comment_notification(existing_ticket, new_comment, sender_user)
                    else:
                        # For external users, use admin as commenter for notification purposes
                        admin_user = User.query.get(primary_admin.id if primary_admin else 1)
                        send_ticket_comment_notification(existing_ticket, new_comment, admin_user)
                    app.logger.info(f"Sent comment notification for ticket #{ticket_id}")
                except Exception as e:
                    app.logger.error(f"Failed to send comment notification: {str(e)}")
                
                return jsonify({
                    'success': True,
                    'ticket_id': existing_ticket.id,
                    'comment_id': new_comment.id,
                    'message': 'Email reply added as comment to existing ticket'
                }), 200
            else:
                app.logger.warning(f"Ticket #{ticket_id} referenced in reply not found, creating new ticket instead")
        
        # Clean up the description - prefer text over HTML, or extract from HTML
        description = text_content.strip() if text_content else ""
        if not description and html_content:
            # Try to extract text from HTML
            import re
            # Remove HTML tags and get clean text
            html_clean = re.sub('<[^<]+?>', '', html_content)
            description = html_clean.strip()
        
        # For forwarded emails, look for original content patterns
        if description and ('forwarded' in description.lower() or 'fwd:' in subject.lower() or 'fw:' in subject.lower()):
            # Try to extract the original message content
            import re
            # Look for common forwarded email patterns
            patterns = [
                r'(?i)---------- forwarded message ----------(.*?)(?=---------- end forwarded message ----------|$)',
                r'(?i)begin forwarded message:(.*?)(?=end forwarded message|$)',
                r'(?i)original message:(.*?)(?=^from:|$)',
                r'(?i)forwarded message:(.*?)(?=^from:|$)'
            ]
            
            for pattern in patterns:
                match = re.search(pattern, description, re.DOTALL)
                if match:
                    original_content = match.group(1).strip()
                    if original_content:
                        description = f"[FORWARDED EMAIL]\nOriginal Content:\n{original_content}\n\n[Full Forward Details]\n{description}"
                    break
        
        # Handle forwarded emails that might have content in different fields
        if not description:
            # Check for other possible content fields from SendGrid
            envelope = email_data.get('envelope', '')
            headers = email_data.get('headers', '')
            attachments = email_data.get('attachments', '0')
            
            # Build description from available metadata
            metadata_parts = []
            if envelope:
                metadata_parts.append(f"Envelope: {envelope}")
            if headers:
                metadata_parts.append(f"Headers: {headers}")
            if attachments != '0':
                metadata_parts.append(f"Attachments: {attachments}")
            
            description = "\n".join(metadata_parts) if metadata_parts else "Email received with no readable content"
        
        # Ensure we have some description
        if not description or description.isspace():
            if 'fwd:' in subject.lower() or 'fw:' in subject.lower():
                description = f"Forwarded email - original content may be in attachments or non-text format\n\nNote: Check email headers and metadata above for routing information."
            else:
                description = f"Email received with no readable content - may contain attachments or be in unsupported format"
        
        # Extract sender name from email
        sender_name = from_email
        if '<' in from_email:
            # Handle "John Doe <john@example.com>" format
            sender_name = from_email.split('<')[0].strip()
            from_email = from_email.split('<')[1].strip('>')
        
        app.logger.info(f"Processing email from: {from_email} (sender name: {sender_name})")
        
        # Try to find existing user by email (case-insensitive)
        sender_user = User.query.filter(User.email.ilike(from_email)).first()
        if sender_user:
            app.logger.info(f"Found existing user for email {from_email}: {sender_user.username}")
        else:
            app.logger.info(f"No existing user found for email {from_email}")
        
        # Determine category - default to first available category
        default_category = TicketCategory.query.first()
        if not default_category:
            app.logger.error("No ticket categories found - cannot create ticket from email")
            return jsonify({'error': 'No ticket categories configured'}), 500
        
        # Generate unique email thread ID for reply tracking
        import uuid
        email_thread_id = f"ticket-{uuid.uuid4().hex[:8]}"
        
        # Create the ticket - assign to sender if known user, otherwise to admin for review
        if sender_user:
            created_by_id = sender_user.id
            external_email = None
            external_name = None
            app.logger.info(f"Creating ticket assigned to sender: {sender_user.username} (ID: {sender_user.id})")
        else:
            # For unknown senders, assign to the primary admin account (admin@obedtv.com)
            primary_admin = User.query.filter_by(email='admin@obedtv.com').first()
            if not primary_admin:
                # Fallback to any admin user if primary admin doesn't exist
                primary_admin = User.query.filter_by(is_admin=True).first()
            created_by_id = primary_admin.id if primary_admin else 1
            external_email = from_email
            external_name = sender_name if sender_name != from_email else None
            app.logger.info(f"Creating ticket assigned to primary admin for unknown sender: {from_email}")
            app.logger.info(f"Primary admin user: {primary_admin.username if primary_admin else 'None'} (ID: {created_by_id})")
        
        app.logger.info(f"FINAL created_by_id that will be used: {created_by_id}")
        
        new_ticket = Ticket(
            title=subject,
            description=f"Email from: {sender_name} ({from_email})\n\n{description}",
            category_id=default_category.id,
            priority=1,  # Medium priority by default
            status='open',
            created_by=created_by_id,
            created_at=datetime.utcnow(),
            external_email=external_email,
            external_name=external_name,
            email_thread_id=email_thread_id,
            email_notifications=True
        )
        
        db.session.add(new_ticket)
        db.session.commit()
        
        app.logger.info(f"Created ticket #{new_ticket.id} from email: {subject}")
        
        # Send confirmation email to ALL users (internal and external)
        try:
            from email_utils import send_email
            confirmation_subject = f"[Ticket #{new_ticket.id}] - {subject}"
            
            # Different email content for internal vs external users
            if sender_user:
                # Internal user - simpler message
                confirmation_body = f"""
                <h3>Ticket Created Successfully</h3>
                <p>Your email has been converted to a support ticket:</p>
                <ul>
                    <li><strong>Ticket #:</strong> {new_ticket.id}</li>
                    <li><strong>Subject:</strong> {subject}</li>
                    <li><strong>Priority:</strong> Medium</li>
                    <li><strong>Status:</strong> Open</li>
                </ul>
                <p>You will receive updates as your ticket is processed.</p>
                """
            else:
                # External user - more detailed message with reply instructions
                confirmation_body = f"""
                <h3>Support Ticket Created</h3>
                <p>Thank you for contacting us. Your email has been converted to a support ticket:</p>
                <ul>
                    <li><strong>Ticket #:</strong> {new_ticket.id}</li>
                    <li><strong>Subject:</strong> {subject}</li>
                    <li><strong>Priority:</strong> Medium</li>
                    <li><strong>Status:</strong> Open</li>
                    <li><strong>Thread ID:</strong> {email_thread_id}</li>
                </ul>
                <p><strong>To reply to this ticket:</strong> Reply to this email and keep the subject line starting with "[Ticket #{new_ticket.id}]"</p>
                <p>You will receive email notifications when your ticket is updated, assigned, or has new comments.</p>
                <p>Our support team will respond to your request as soon as possible.</p>
                """
            
            send_email([from_email], confirmation_subject, confirmation_body)
            app.logger.info(f"Sent confirmation email to {from_email} ({'internal' if sender_user else 'external'} user)")
        except Exception as e:
            app.logger.error(f"Failed to send confirmation email: {str(e)}")
        
        return jsonify({
            'success': True,
            'ticket_id': new_ticket.id,
            'message': 'Email converted to ticket successfully'
        }), 200
        
    except Exception as e:
        app.logger.error(f"Error processing inbound email: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to process email',
            'message': str(e)
        }), 500

@app.route('/api/test-email-webhook', methods=['GET'])
@login_required
def test_email_webhook():
    """Test endpoint to verify email webhook functionality"""
    if not current_user.is_admin:
        return jsonify({'error': 'Admin access required'}), 403
    
    # Simulate a test email
    test_data = {
        'from': 'test@example.com',
        'subject': 'Test Email Ticket',
        'text': 'This is a test email to verify the email-to-ticket system is working.',
        'to': 'tickets@mail1.opscal.io'
    }
    
    app.logger.info("Testing email webhook with simulated data")
    
    # Simulate the webhook processing
    try:
        description = test_data['text']
        sender_name = test_data['from']
        subject = test_data['subject']
        
        # Get default category
        default_category = TicketCategory.query.first()
        if not default_category:
            return jsonify({'error': 'No ticket categories found'}), 500
        
        # Create test ticket
        test_ticket = Ticket(
            title=f"[TEST] {subject}",
            description=f"Email from: {sender_name}\n\n{description}",
            category_id=default_category.id,
            priority=1,
            status='open',
            created_by=current_user.id,
            created_at=datetime.utcnow()
        )
        
        db.session.add(test_ticket)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'ticket_id': test_ticket.id,
            'message': 'Test ticket created successfully',
            'webhook_url': url_for('inbound_email_webhook', _external=True)
        })
        
    except Exception as e:
        return jsonify({
            'error': 'Test failed',
            'message': str(e)
        }), 500

@app.route('/login', methods=['GET', 'POST'])
def login():
    app.logger.debug(f"Login attempt - Method: {request.method}")
    app.logger.debug(f"Session before login: {session}")

    if current_user.is_authenticated:
        app.logger.debug(f"Already authenticated user: {current_user.username}")
        app.logger.debug(f"Current session: {session}")
        return redirect(url_for('calendar'))

    form = LoginForm()
    if form.validate_on_submit():
        app.logger.debug(f"Login form submitted for email: {form.email.data}")
        user = User.query.filter_by(email=form.email.data).first()
        if user and user.check_password(form.password.data):
            # Set session to be permanent (7 days)
            session.permanent = True
            login_user(user, remember=form.remember_me.data)
            app.logger.info(f"User {user.username} logged in successfully")
            app.logger.debug(f"Session after login: {session}")
            next_page = request.args.get('next')
            app.logger.debug(f"Redirecting to: {next_page if next_page else 'calendar'}")
            return redirect(next_page if next_page else url_for('calendar'))
        app.logger.warning(f"Failed login attempt for email: {form.email.data}")
        flash('Invalid email or password')
    return render_template('login.html', form=form)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('calendar'))

    form = RegistrationForm()
    if form.validate_on_submit():
        user = User(username=form.username.data, email=form.email.data)
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash('Registration successful! Please log in.')
        return redirect(url_for('login'))
    return render_template('register.html', form=form)

# Logout route is handled in auth.py - removed duplicate to prevent conflicts

@app.route('/api/active_users')
@login_required
def get_active_users():
    """Get users who have schedules active at the current time"""
    # The @login_required decorator will handle redirecting unauthenticated users
    # But we'll add an explicit check for API JSON responses
    if not current_user.is_authenticated:
        app.logger.debug("Unauthorized access to path: /api/active_users")
        return jsonify({'error': 'Authentication required'}), 401

    try:
        # Get current time in UTC since our database stores times in UTC
        current_time = datetime.now(pytz.UTC)
        app.logger.debug(f"Current time (UTC): {current_time}")
        
        # Find who is actively scheduled right now
        active_users_query = (db.session.query(
                User, Schedule, Location
            )
            .join(Schedule, User.id == Schedule.technician_id)
            .outerjoin(Location, Schedule.location_id == Location.id)
            .filter(
                Schedule.start_time <= current_time,
                Schedule.end_time > current_time,
                ~Schedule.time_off  # Exclude time off entries
            ))
            
        app.logger.debug(f"Active users query SQL: {str(active_users_query)}")
        active_users_data = active_users_query.all()
        app.logger.debug(f"Found {len(active_users_data)} active schedules")

        # Build the result - only show currently active users
        result = []
        
        # Process active users data
        user_tz = current_user.get_timezone_obj()
        for user, schedule, location in active_users_data:
            # Skip if any key component is None
            if not user or not schedule:
                app.logger.warning(f"Skipping schedule with missing user or schedule data")
                continue
                
            try:
                # Convert schedule times to user's timezone
                start_time = schedule.start_time.astimezone(user_tz)
                end_time = schedule.end_time.astimezone(user_tz)
                
                # Add this active user to the result
                result.append({
                    'username': user.username,
                    'color': user.color,
                    'profile_picture': user.profile_picture,
                    'schedule': {
                        'start_time': start_time.strftime('%H:%M'),
                        'end_time': end_time.strftime('%H:%M'),
                        'description': schedule.description or ''
                    },
                    'location': {
                        'name': location.name if location else 'No Location',
                        'description': location.description if location else ''
                    }
                })
            except Exception as inner_e:
                app.logger.error(f"Error processing schedule for user {user.id}: {str(inner_e)}")
                # Continue processing other users

        app.logger.debug(f"Returning {len(result)} active users")
        return jsonify(result)
    except Exception as e:
        app.logger.error(f"Error in get_active_users: {str(e)}")
        app.logger.exception("Detailed exception information:")
        return jsonify([]), 200  # Return empty array instead of error

@app.route('/profile')
@login_required
def profile():
    form = EditUserForm(obj=current_user)
    password_form = ChangePasswordForm()
    
    # For debugging
    print(f"is_mobile_device() in profile: {is_mobile_device()}")
    is_mobile = is_mobile_device()  # Force evaluation
    print(f"is_mobile value in profile: {is_mobile}")
    
    if is_mobile_device():
        # Use mobile template with timezone list
        return render_template('mobile_profile.html', 
                             form=form, 
                             password_form=password_form,
                             timezones=pytz.common_timezones)
    
    # Use desktop template
    return render_template('profile.html', form=form, password_form=password_form)

@app.route('/profile/update', methods=['POST'])
@login_required
def update_profile():
    color = request.form.get('color')
    if color:
        try:
            current_user.color = color
            db.session.commit()
            flash('Profile updated successfully!')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating profile: {str(e)}")
            flash('Error updating profile. Please try again.')
    return redirect(url_for('profile'))

@app.route('/profile/change-password', methods=['POST'])
@login_required
def change_password():
    form = ChangePasswordForm()
    if form.validate_on_submit():
        if current_user.check_password(form.current_password.data):
            current_user.set_password(form.new_password.data)
            db.session.commit()
            flash('Password updated successfully!')
            return redirect(url_for('profile'))
        flash('Current password is incorrect')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'{getattr(form, field).label.text}: {error}')
    return redirect(url_for('profile'))


@app.route('/toggle-theme', methods=['POST'])
@login_required
def toggle_theme():
    try:
        # Toggle between light and dark theme
        current_theme = current_user.theme_preference or 'dark'  # Default to dark if None
        new_theme = 'light' if current_theme == 'dark' else 'dark'
        
        # Update user preference
        current_user.theme_preference = new_theme
        db.session.commit()
        
        flash(f'Theme updated to {new_theme} mode')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error toggling theme: {str(e)}")
        flash('Error updating theme preference. Please try again.')
    
    # Redirect back to the page they came from or default to profile
    return redirect(request.referrer or url_for('profile'))


@app.route('/admin/locations', methods=['GET', 'POST'])
@login_required
def admin_locations():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    form = LocationForm()
    if form.validate_on_submit():
        try:
            location = Location(
                name=form.name.data,
                description=form.description.data,
                active=form.active.data
            )
            db.session.add(location)
            db.session.commit()
            flash('Location added successfully!')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error creating location: {str(e)}")
            flash('Error creating location. Please try again.')

    locations = Location.query.order_by(Location.name).all()
    return render_template('admin/locations.html', locations=locations, form=form)

@app.route('/calendar')
@login_required
def calendar():
    week_start = request.args.get('week_start')
    location_filter = request.args.get('location_id', type=int)
    
    if week_start:
        week_start = datetime.strptime(week_start, '%Y-%m-%d')
        week_start = current_user.get_timezone_obj().localize(
            week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        )
    else:
        week_start = datetime.now(current_user.get_timezone_obj())
        week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start -= timedelta(days=week_start.weekday())

    # Convert to UTC for database query
    week_start_utc = week_start.astimezone(pytz.UTC)
    week_end_utc = (week_start + timedelta(days=7)).astimezone(pytz.UTC)

    # Query schedules in UTC with optional location filter
    # Use overlap logic: schedule overlaps with week if it starts before week ends and ends after week starts
    query = Schedule.query.filter(
        Schedule.start_time < week_end_utc,
        Schedule.end_time > week_start_utc
    )

    if location_filter:
        query = query.filter(Schedule.location_id == location_filter)
        
    # Order by start_time to ensure chronological display
    query = query.order_by(Schedule.start_time)

    schedules = query.all()
    
    # Debug logging
    app.logger.debug(f"Calendar query found {len(schedules)} schedules")
    app.logger.debug(f"Week start UTC: {week_start_utc}")
    app.logger.debug(f"Week end UTC: {week_end_utc}")
    
    # Count time-off entries
    time_off_count = sum(1 for s in schedules if s.time_off)
    all_day_time_off_count = sum(1 for s in schedules if s.time_off and s.all_day)
    app.logger.debug(f"Found {time_off_count} time-off entries, {all_day_time_off_count} all-day")
    
    # Debug: check if testuser schedule is in the list
    testuser_schedules = [s for s in schedules if s.technician.username == 'testuser']
    app.logger.debug(f"Found {len(testuser_schedules)} testuser schedules")
    for ts in testuser_schedules:
        app.logger.debug(f"  Testuser schedule {ts.id}: {ts.start_time} to {ts.end_time}, time_off={ts.time_off}, all_day={ts.all_day}")
    


    # Convert schedule times to user's timezone
    user_tz = current_user.get_timezone_obj()
    for schedule in schedules:
        if schedule.start_time.tzinfo is None:
            schedule.start_time = pytz.UTC.localize(schedule.start_time)
        if schedule.end_time.tzinfo is None:
            schedule.end_time = pytz.UTC.localize(schedule.end_time)

        # Special handling for all-day time-off events to prevent timezone date shifting
        if schedule.time_off and schedule.all_day:
            # For all-day events, we need to determine the intended calendar date
            # Since existing entries were created in Chicago time, we reverse-engineer the date
            utc_time = schedule.start_time.astimezone(pytz.UTC)
            
            # Try to determine the original calendar date by checking common US timezones
            chicago_tz = pytz.timezone('America/Chicago')
            pacific_tz = pytz.timezone('America/Los_Angeles')
            
            chicago_display = utc_time.astimezone(chicago_tz)
            pacific_display = utc_time.astimezone(pacific_tz)
            
            # If the UTC time matches Chicago midnight conversion pattern, use Chicago date
            chicago_midnight = chicago_tz.localize(datetime.combine(chicago_display.date(), time(0, 0)))
            if utc_time == chicago_midnight.astimezone(pytz.UTC):
                intended_date = chicago_display.date()
                app.logger.debug(f"All-day OOO {schedule.id}: Detected Chicago-created entry for {intended_date}")
            else:
                # Otherwise, use the current user's timezone date
                intended_date = utc_time.astimezone(user_tz).date()
                app.logger.debug(f"All-day OOO {schedule.id}: Using user timezone date {intended_date}")
            
            # Display as all-day in user's timezone for the intended date
            schedule.start_time = user_tz.localize(datetime.combine(intended_date, time(0, 0)))
            schedule.end_time = user_tz.localize(datetime.combine(intended_date, time(23, 59)))
            app.logger.debug(f"All-day display fix for schedule {schedule.id}: {intended_date} → {schedule.start_time} to {schedule.end_time}")
        else:
            schedule.start_time = schedule.start_time.astimezone(user_tz)
            schedule.end_time = schedule.end_time.astimezone(user_tz)

    form = ScheduleForm()
    if current_user.is_admin:
        form.technician.choices = [(u.id, u.username) for u in User.query.all()]
    else:
        form.technician.choices = [(current_user.id, current_user.username)]
        form.technician.data = current_user.id

    # Add location choices to the form
    form.location_id.choices = [(l.id, l.name) for l in Location.query.filter_by(active=True).order_by(Location.name).all()]

    # Get all active locations for the filter dropdown
    locations = Location.query.filter_by(active=True).order_by(Location.name).all()

    # Debug the mobile detection
    print(f"is_mobile_device() in calendar: {is_mobile_device()}")
    is_mobile = is_mobile_device()  # Force evaluation
    print(f"is_mobile value: {is_mobile}")
    
    # Check for force mobile parameter for testing
    force_mobile = request.args.get('mobile') == 'true'
    
    # Check if user is on a mobile device
    if is_mobile_device() or force_mobile:
        print("Using mobile template for calendar")
        return render_template('mobile_calendar.html', 
                            schedules=schedules,
                            week_start=week_start,
                            week_end=week_start + timedelta(days=7),
                            form=form,
                            locations=locations,
                            selected_location=location_filter,
                            today=datetime.now(current_user.get_timezone_obj()),
                            user_timezone=str(current_user.get_timezone_obj()),
                            datetime=datetime,
                            timedelta=timedelta)
    else:
        return render_template('calendar.html', 
                            schedules=schedules,
                            week_start=week_start,
                            week_end=week_start + timedelta(days=7),
                            form=form,
                            locations=locations,
                            selected_location=location_filter,
                            today=datetime.now(current_user.get_timezone_obj()),
                            user_timezone=str(current_user.get_timezone_obj()),
                            datetime=datetime,
                            timedelta=timedelta)

@app.route('/schedule/new', methods=['GET', 'POST'])
@login_required
def new_schedule():
    # Get the week_start parameter to maintain the same view
    week_start = request.args.get('week_start') or request.form.get('week_start')
    personal_view = request.args.get('personal_view') == 'true' or request.form.get('personal_view') == 'true'
    
    form = ScheduleForm()

    # Set up technician choices
    if current_user.is_admin:
        form.technician.choices = [(u.id, u.username) for u in User.query.all()]
    else:
        form.technician.choices = [(current_user.id, current_user.username)]
        form.technician.data = current_user.id

    # Set up location choices
    locations = Location.query.filter_by(active=True).order_by(Location.name).all()
    form.location_id.choices = [(l.id, l.name) for l in locations]
    # Add an empty choice if no locations exist
    if not locations:
        form.location_id.choices = [(0, 'No locations available')]

    # Check if we're dealing with a mobile form submission (with schedule_date, start_hour, end_hour)
    schedule_date = request.form.get('schedule_date')
    start_hour = request.form.get('start_hour')
    end_hour = request.form.get('end_hour')
    
    # Debug log to see what's being submitted
    app.logger.debug(f"Form submission: {request.form}")
    
    is_mobile_submission = schedule_date and start_hour and end_hour
    
    if is_mobile_submission:
        # We're getting data from the mobile form
        app.logger.debug(f"Mobile form data received: date={schedule_date}, start={start_hour}, end={end_hour}")
        try:
            # Parse the date
            date_obj = datetime.strptime(schedule_date, '%Y-%m-%d').date()
            
            # Parse start_hour with minutes (e.g., "14:30" or "14:00")
            if ':' in start_hour:
                start_hour_parts = start_hour.split(':')
                start_hour_int = int(start_hour_parts[0])
                start_minute_int = int(start_hour_parts[1]) if len(start_hour_parts) > 1 else 0
            else:
                # Legacy format: just hour
                start_hour_int = int(start_hour)
                start_minute_int = 0
            start_time_obj = datetime.combine(date_obj, time(hour=start_hour_int, minute=start_minute_int))
            
            # Parse end_hour with minutes (e.g., "16:30" or "00:00")
            if ':' in end_hour:
                end_hour_parts = end_hour.split(':')
                end_hour_int = int(end_hour_parts[0])
                end_minute_int = int(end_hour_parts[1]) if len(end_hour_parts) > 1 else 0
            else:
                # Legacy format: just hour
                end_hour_int = int(end_hour)
                end_minute_int = 0
                
            # If end hour is 0 (midnight), it should be the next day
            if end_hour_int == 0:
                # Set to midnight of the same day, will add a day later for UTC
                end_time_obj = datetime.combine(date_obj, time(hour=0, minute=end_minute_int))
            else:
                end_time_obj = datetime.combine(date_obj, time(hour=end_hour_int, minute=end_minute_int))
            
            # Assign to form
            form.start_time.data = start_time_obj
            form.end_time.data = end_time_obj
            
            # Handle form fields
            form.description.data = request.form.get('description', '')
            
            # Handle time_off checkbox
            time_off_val = request.form.get('time_off')
            form.time_off.data = bool(time_off_val == 'on' or time_off_val == 'true' or time_off_val == '1')
            
            # Handle all_day checkbox  
            all_day_val = request.form.get('all_day')
            form.all_day.data = bool(all_day_val == 'on' or all_day_val == 'true' or all_day_val == '1')
            
            # If OOO (all_day) is checked, automatically set times to cover the full day
            if form.all_day.data:
                # Override the start and end times to cover the full day
                form.start_time.data = datetime.combine(date_obj, time(0, 0))  # 00:00
                form.end_time.data = datetime.combine(date_obj, time(23, 0))  # 23:00 (matches mobile form)
                app.logger.debug(f"OOO checked: automatically set times to 00:00-23:00 for {date_obj}")
            
            # Get the location_id if present
            location_id = request.form.get('location_id')
            if location_id and location_id.isdigit():
                form.location_id.data = int(location_id)
            else:
                form.location_id.data = 0
                
            # Handle repeat days for mobile form
            # First try the new mini-calendar format (comma-separated list in a single field)
            repeat_days_list_field = request.form.get('repeat_days_list')
            if repeat_days_list_field:
                form.repeat_days.data = repeat_days_list_field
                app.logger.debug(f"Mobile form (mini-calendar): Repeat days selected: {form.repeat_days.data}")
            else:
                # Fall back to original checkbox format
                repeat_days_list = request.form.getlist('repeat_days')
                if repeat_days_list:
                    form.repeat_days.data = ','.join(repeat_days_list)
                    app.logger.debug(f"Mobile form (checkboxes): Repeat days selected: {form.repeat_days.data}")
                
            # Mobile validation successful
            is_mobile_validation_successful = True
            
        except Exception as e:
            app.logger.error(f"Error parsing mobile form data: {str(e)}")
            flash('Invalid date or time format. Please try again.')
            return redirect(url_for('calendar', week_start=week_start))

    if form.validate_on_submit() or is_mobile_submission:
        try:
            app.logger.debug(f"Processing form data: {request.form}")
            schedule_id = request.form.get('schedule_id')
            technician_id = form.technician.data if current_user.is_admin else current_user.id
            
            # Also check if we have a week_start in the form
            if not week_start:
                week_start = request.form.get('week_start')

            user_tz = current_user.get_timezone_obj()
            start_time = user_tz.localize(form.start_time.data)
            end_time = user_tz.localize(form.end_time.data)

            start_time_utc = start_time.astimezone(pytz.UTC)
            end_time_utc = end_time.astimezone(pytz.UTC)

            if end_time.hour == 0 and end_time.minute == 0:
                end_time_utc = end_time_utc + timedelta(days=1)

            if end_time_utc <= start_time_utc:
                flash('End time must be after start time.')
                return redirect(url_for('calendar', week_start=week_start))

            overlapping_query = Schedule.query.filter(
                Schedule.technician_id == technician_id,
                Schedule.id != (int(schedule_id) if schedule_id else None),
                Schedule.start_time < end_time_utc,
                Schedule.end_time > start_time_utc
            )

            overlapping_schedules = overlapping_query.first()

            if overlapping_schedules and not form.time_off.data:
                flash('Schedule conflicts with existing appointments.')
                if personal_view:
                    return redirect(url_for('personal_schedule', week_start=week_start))
                else:
                    return redirect(url_for('calendar', week_start=week_start))

            # Check if we have repeat days selected from any of the possible sources
            repeat_days = None
            
            # First check for direct repeat days list from our direct selector (comma-separated string)
            direct_repeat_days = request.form.get('direct_repeat_days_list')
            if direct_repeat_days:
                app.logger.debug(f"Found direct_repeat_days_list: {direct_repeat_days}")
                repeat_days = direct_repeat_days
                
            # Next check the repeat_days hidden field (should be the same as direct_repeat_days_list)
            if not repeat_days:
                repeat_days_field = request.form.get('repeat_days')
                if repeat_days_field:
                    app.logger.debug(f"Found repeat_days hidden field: {repeat_days_field}")
                    repeat_days = repeat_days_field
            
            # Fallback to checkbox approach used by the old form system
            if not repeat_days:
                app.logger.debug("Checking for repeat_days checkboxes")
                repeat_days_checkboxes = request.form.getlist('repeat_days')
                if repeat_days_checkboxes:
                    app.logger.debug(f"Found repeat_days checkboxes: {repeat_days_checkboxes}")
                    repeat_days = ','.join(repeat_days_checkboxes)
            
            # Fallback to the WTForms field data
            if not repeat_days:
                app.logger.debug("No repeat_days values found in request, checking form.repeat_days.data")
                if form.repeat_days and form.repeat_days.data:
                    repeat_days = form.repeat_days.data
                    
            app.logger.debug(f"Final repeat_days value: {repeat_days}")
            
            if schedule_id:
                # Editing an existing schedule - doesn't support multi-day editing
                schedule = Schedule.query.get_or_404(schedule_id)
                if schedule.technician_id != current_user.id and not current_user.is_admin:
                    flash('You do not have permission to edit this schedule.')
                    if personal_view:
                        return redirect(url_for('personal_schedule', week_start=week_start))
                    else:
                        return redirect(url_for('calendar', week_start=week_start))

                old_desc = schedule.description
                
                # Check if this is an all-day event for editing
                is_all_day_edit = getattr(form, 'all_day', None) and form.all_day.data
                
                # Auto-set time_off to True when all_day is checked (simplified vacation system)
                if is_all_day_edit:
                    form.time_off.data = True
                
                if is_all_day_edit and form.time_off.data:
                    # For all-day time-off events during editing, create them in the user's local timezone
                    # then convert to UTC for database storage
                    user_tz = current_user.get_timezone_obj()
                    local_date = start_time.date()
                    
                    app.logger.debug(f"All-day OOO edit debug:")
                    app.logger.debug(f"  User timezone: {user_tz}")
                    app.logger.debug(f"  Local date: {local_date}")
                    app.logger.debug(f"  Original start_time: {start_time}")
                    
                    # Create all-day times in user's local timezone (00:00-23:59)
                    start_time_local = user_tz.localize(datetime.combine(local_date, time(0, 0)))
                    end_time_local = user_tz.localize(datetime.combine(local_date, time(23, 59)))
                    
                    app.logger.debug(f"  Local times: {start_time_local} to {end_time_local}")
                    
                    # Convert to UTC for database storage
                    start_time_utc = start_time_local.astimezone(pytz.UTC)
                    end_time_utc = end_time_local.astimezone(pytz.UTC)
                    
                    app.logger.debug(f"  UTC times for storage: {start_time_utc} to {end_time_utc}")
                
                schedule.start_time = start_time_utc
                schedule.end_time = end_time_utc
                schedule.description = form.description.data
                schedule.time_off = form.time_off.data
                schedule.all_day = is_all_day_edit
                schedule.location_id = form.location_id.data if form.location_id.data != 0 else None
                if current_user.is_admin:
                    schedule.technician_id = technician_id
                send_schedule_notification(schedule, 'updated', f"Schedule updated by {current_user.username}")
                
                db.session.commit()
                flash('Schedule updated successfully!')
                
            else:
                # Creating new schedule(s)
                schedules_created = 0
                
                # First, always create a schedule for the primary date
                app.logger.debug(f"Creating primary schedule for date: {schedule_date}")
                
                # Check if this is an all-day event and handle timezone storage properly
                is_all_day = getattr(form, 'all_day', None) and form.all_day.data
                
                # Auto-set time_off to True when all_day is checked (simplified vacation system)
                if is_all_day:
                    form.time_off.data = True
                
                if is_all_day and form.time_off.data:
                    # For all-day time-off events, create them in the user's local timezone
                    # then convert to UTC for database storage
                    user_tz = current_user.get_timezone_obj()
                    local_date = start_time.date()
                    
                    app.logger.debug(f"All-day OOO creation debug:")
                    app.logger.debug(f"  User timezone: {user_tz}")
                    app.logger.debug(f"  Local date: {local_date}")
                    app.logger.debug(f"  Original start_time: {start_time}")
                    
                    # Create all-day times in user's local timezone (00:00-23:59)
                    start_time_local = user_tz.localize(datetime.combine(local_date, time(0, 0)))
                    end_time_local = user_tz.localize(datetime.combine(local_date, time(23, 59)))
                    
                    app.logger.debug(f"  Local times: {start_time_local} to {end_time_local}")
                    
                    # Convert to UTC for database storage
                    start_time_utc = start_time_local.astimezone(pytz.UTC)
                    end_time_utc = end_time_local.astimezone(pytz.UTC)
                    
                    app.logger.debug(f"  UTC times for storage: {start_time_utc} to {end_time_utc}")
                    app.logger.debug(f"All-day vacation created: Local {start_time_local} to {end_time_local}, UTC {start_time_utc} to {end_time_utc}")
                else:
                    # Use the originally calculated UTC times for regular schedules
                    app.logger.debug(f"Regular schedule created: {start_time_utc} to {end_time_utc}")

                # Check for existing all-day time-off entry before creating new one
                existing_all_day_entry = None
                if is_all_day and form.time_off.data:
                    # Check if there's already an all-day time-off entry for this technician on this date
                    local_date = start_time.date()
                    date_start = pytz.UTC.localize(datetime.combine(local_date, time(0, 0)))
                    date_end = pytz.UTC.localize(datetime.combine(local_date, time(23, 59, 59)))
                    
                    existing_all_day_entry = Schedule.query.filter(
                        Schedule.technician_id == technician_id,
                        Schedule.time_off == True,
                        Schedule.all_day == True,
                        Schedule.start_time >= date_start,
                        Schedule.start_time <= date_end
                    ).first()
                    
                    if existing_all_day_entry:
                        app.logger.debug(f"Found existing all-day time-off entry for {schedule_date}, updating instead of creating new")
                        # Update existing entry instead of creating new one
                        existing_all_day_entry.start_time = start_time_utc
                        existing_all_day_entry.end_time = end_time_utc
                        existing_all_day_entry.description = form.description.data
                        existing_all_day_entry.location_id = form.location_id.data if form.location_id.data != 0 else None
                        schedules_created += 1
                        app.logger.debug(f"Updated existing all-day time-off entry for {schedule_date}")
                    
                    # Handle OOO conflict prevention: remove or gray out existing regular schedules for this day
                    conflicting_schedules = Schedule.query.filter(
                        Schedule.technician_id == technician_id,
                        Schedule.id != (existing_all_day_entry.id if existing_all_day_entry else None),
                        Schedule.time_off == False,  # Only affect regular work schedules
                        Schedule.start_time >= date_start,
                        Schedule.end_time <= date_end + timedelta(hours=23, minutes=59)
                    ).all()
                    
                    if conflicting_schedules:
                        app.logger.debug(f"Found {len(conflicting_schedules)} conflicting regular schedules for OOO on {schedule_date}")
                        for conflict_schedule in conflicting_schedules:
                            # Mark as cancelled/grayed out by adding OOO note to description
                            original_desc = conflict_schedule.description or ""
                            if "CANCELLED - OOO" not in original_desc:
                                conflict_schedule.description = f"CANCELLED - OOO: {original_desc}".strip()
                                app.logger.debug(f"Marked schedule {conflict_schedule.id} as cancelled due to OOO")
                        
                        flash(f'OOO created. {len(conflicting_schedules)} existing schedule(s) marked as cancelled.', 'info')
                
                # Create primary schedule only if no existing all-day entry was found
                if not existing_all_day_entry:
                    primary_schedule = Schedule(
                        technician_id=technician_id,
                        start_time=start_time_utc,
                        end_time=end_time_utc,
                        description=form.description.data,
                        time_off=form.time_off.data,
                        all_day=is_all_day,
                        location_id=form.location_id.data if form.location_id.data != 0 else None
                    )
                    db.session.add(primary_schedule)
                    schedules_created += 1
                    app.logger.debug(f"Primary schedule created for {schedule_date}")
                
                # Then handle additional dates if repeat_days is provided
                if repeat_days:
                    # Multi-day scheduling for additional dates
                    try:
                        # Split the comma-separated string of dates
                        if isinstance(repeat_days, str):
                            dates = [d.strip() for d in repeat_days.split(',') if d.strip()]
                            app.logger.debug(f"Processing additional schedules from string. Dates: {dates}")
                        elif isinstance(repeat_days, list):
                            dates = repeat_days
                            app.logger.debug(f"Processing additional schedules from list. Dates: {dates}")
                        else:
                            app.logger.warning(f"Unexpected repeat_days type: {type(repeat_days)}")
                            dates = []
                        
                        # Get the primary date string for filtering
                        primary_date_str = schedule_date
                        if not primary_date_str:
                            # If no schedule_date provided (e.g., in desktop version), extract from start_time
                            try:
                                if form.start_time and form.start_time.data:
                                    primary_date_str = form.start_time.data.strftime('%Y-%m-%d')
                                else:
                                    app.logger.warning("No start_time data available in form")
                                    primary_date_str = datetime.now().strftime('%Y-%m-%d')
                            except Exception as e:
                                app.logger.error(f"Error getting date from form.start_time.data: {e}")
                                # Fallback to current date
                                primary_date_str = datetime.now().strftime('%Y-%m-%d')
                        elif not isinstance(primary_date_str, str):
                            try:
                                primary_date_str = primary_date_str.strftime('%Y-%m-%d')
                            except Exception as e:
                                app.logger.error(f"Error formatting primary_date_str: {e}")
                                primary_date_str = datetime.now().strftime('%Y-%m-%d')
                        app.logger.debug(f"Primary date (to exclude from additional dates): {primary_date_str}")
                        
                        # Filter out the primary date since we already created a schedule for it
                        dates = [date for date in dates if date.strip() != primary_date_str]
                        app.logger.debug(f"Additional dates list after removing primary date: {dates}")
                        
                        # Validate all dates have the proper format
                        valid_dates = []
                        for date_str in dates:
                            try:
                                # Check if this is a valid date string
                                date_obj = datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
                                valid_dates.append(date_str.strip())
                            except ValueError:
                                app.logger.warning(f"Invalid date format ignored: {date_str}")
                                
                        dates = valid_dates
                        app.logger.debug(f"Valid dates to process: {dates}")
                        
                    except Exception as e:
                        app.logger.error(f"Error processing repeat days: {str(e)}")
                        import traceback
                        app.logger.error(f"Repeat days traceback: {traceback.format_exc()}")
                        app.logger.error(f"Original repeat_days value: {repeat_days}")
                        # Initialize empty dates list to avoid errors
                        dates = []
                    
                    if not dates:
                        app.logger.debug("No additional dates selected besides primary date")
                        # We already created the primary schedule, so continue processing
                    
                    # Create a schedule for each selected day
                    for date_str in dates:
                        # Parse the date
                        day_date = datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
                        
                        # Handle all-day events for multi-day schedules
                        if is_all_day and form.time_off.data:
                            # For all-day time-off events, store as 00:00 to 23:59 in the user's local date
                            day_start_time = user_tz.localize(datetime.combine(day_date, time(0, 0)))
                            day_end_time = user_tz.localize(datetime.combine(day_date, time(23, 59)))
                        else:
                            # Create a new datetime using the date from day_date and time from start_time/end_time
                            day_start_time = user_tz.localize(
                                datetime.combine(day_date, datetime.min.time().replace(
                                    hour=start_time.hour, minute=start_time.minute
                                ))
                            )
                            day_end_time = user_tz.localize(
                                datetime.combine(day_date, datetime.min.time().replace(
                                    hour=end_time.hour, minute=end_time.minute
                                ))
                            )
                        
                        # Convert to UTC
                        day_start_time_utc = day_start_time.astimezone(pytz.UTC)
                        day_end_time_utc = day_end_time.astimezone(pytz.UTC)
                        
                        # Handle midnight end time (only for non-all-day events)
                        if not is_all_day and end_time.hour == 0 and end_time.minute == 0:
                            day_end_time_utc = day_end_time_utc + timedelta(days=1)
                            
                        # Check for overlapping schedules for this specific day
                        overlapping_query = Schedule.query.filter(
                            Schedule.technician_id == technician_id,
                            Schedule.start_time < day_end_time_utc,
                            Schedule.end_time > day_start_time_utc
                        )
                        overlapping_schedule = overlapping_query.first()
                        
                        if overlapping_schedule and not form.time_off.data:
                            app.logger.warning(f"Skipping schedule for {date_str} due to conflict")
                            continue
                            
                        # Create the schedule for this day
                        schedule = Schedule(
                            technician_id=technician_id,
                            start_time=day_start_time_utc,
                            end_time=day_end_time_utc,
                            description=form.description.data,
                            time_off=form.time_off.data,
                            all_day=is_all_day,
                            location_id=form.location_id.data if form.location_id.data != 0 else None
                        )
                        db.session.add(schedule)
                        schedules_created += 1
                        
                    if schedules_created > 0:
                        send_schedule_notification(primary_schedule, 'created', 
                            f"Multiple schedules created by {current_user.username}")
                        db.session.commit()
                        flash(f'{schedules_created} schedules created successfully!')
                    else:
                        # We already created the primary schedule, so we should still commit
                        send_schedule_notification(primary_schedule, 'created', 
                            f"Schedule created by {current_user.username}")
                        db.session.commit()
                        flash('Schedule created for the primary date.')
                        
                else:
                    # Single day scheduling - primary schedule already created above
                    send_schedule_notification(primary_schedule, 'created', f"Schedule created by {current_user.username}")
                    db.session.commit()
                    flash('Schedule created successfully!')
            if personal_view:
                return redirect(url_for('personal_schedule', week_start=week_start))
            else:
                return redirect(url_for('calendar', week_start=week_start))

        except Exception as e:
            db.session.rollback()
            flash('Error saving schedule. Please check the time entries.')
            app.logger.error(f"Error saving schedule: {str(e)}")
            # Add detailed traceback for debugging
            import traceback
            app.logger.error(f"Traceback: {traceback.format_exc()}")
            # Log the request form data for debugging
            app.logger.error(f"Form data: {request.form}")
            if personal_view:
                return redirect(url_for('personal_schedule', week_start=week_start))
            else:
                return redirect(url_for('calendar', week_start=week_start))

    if personal_view:
        return redirect(url_for('personal_schedule', week_start=week_start))
    else:
        return redirect(url_for('calendar', week_start=week_start))

@app.route('/schedule/delete/<int:schedule_id>')
@login_required
def delete_schedule(schedule_id):
    # Debug information
    app.logger.debug(f"Delete schedule request for schedule_id: {schedule_id}")
    app.logger.debug(f"Request args: {request.args}")
    app.logger.debug(f"Request form: {request.form}")
    
    # Get current week start to maintain the same view
    week_start = request.args.get('week_start') or request.form.get('week_start')
    app.logger.debug(f"Week start: {week_start}")
    
    # Check both personal_view and return_to parameters for backward compatibility
    personal_view = (request.args.get('personal_view') == 'true' or 
                     request.form.get('personal_view') == 'true' or 
                     request.args.get('return_to') == 'personal_schedule' or 
                     request.form.get('return_to') == 'personal_schedule')
    app.logger.debug(f"Personal view: {personal_view}")
    
    # Try to get the schedule, but don't force a 404 if not found
    schedule = Schedule.query.get(schedule_id)
    if not schedule:
        app.logger.warning(f"Schedule with ID {schedule_id} not found")
        flash('Schedule not found or already deleted.')
        if personal_view:
            return redirect(url_for('personal_schedule', week_start=week_start))
        else:
            return redirect(url_for('calendar', week_start=week_start))
    
    app.logger.debug(f"Schedule found: {schedule.id}, technician_id: {schedule.technician_id}")

    if schedule.technician_id != current_user.id and not current_user.is_admin:
        flash('You do not have permission to delete this schedule.')
        if personal_view:
            return redirect(url_for('personal_schedule', week_start=week_start))
        else:
            return redirect(url_for('calendar', week_start=week_start))

    try:
        send_schedule_notification(schedule, 'deleted', f"Schedule deleted by {current_user.username}")
        db.session.delete(schedule)
        db.session.commit()
        flash('Schedule deleted successfully!')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting schedule.')
        app.logger.error(f"Error deleting schedule: {str(e)}")

    # Redirect back to the same week view
    if personal_view:
        return redirect(url_for('personal_schedule', week_start=week_start))
    else:
        return redirect(url_for('calendar', week_start=week_start))

@app.route('/schedule/copy_previous_week', methods=['POST'])
@login_required
def copy_previous_week_schedules():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    try:
        if not request.form.get('csrf_token') or not request.form.get('csrf_token') == request.form.get('csrf_token'):
            flash('Invalid request. Please try again.')
            return redirect(url_for('calendar'))

        target_week_start_str = request.form.get('target_week_start')
        if not target_week_start_str:
            flash('Invalid week start date.')
            return redirect(url_for('calendar'))

        user_tz = current_user.get_timezone_obj()

        # Convert target week start to user's timezone
        target_week_start = user_tz.localize(
            datetime.strptime(target_week_start_str, '%Y-%m-%d')
        )
        previous_week_start = target_week_start - timedelta(days=7)

        # Convert to UTC for database operations
        target_week_start_utc = target_week_start.astimezone(pytz.UTC)
        target_week_end_utc = (target_week_start + timedelta(days=7)).astimezone(pytz.UTC)
        previous_week_start_utc = previous_week_start.astimezone(pytz.UTC)
        previous_week_end_utc = (previous_week_start + timedelta(days=7)).astimezone(pytz.UTC)

        app.logger.debug(f"Copying schedules from {previous_week_start_utc} to {previous_week_end_utc}")
        app.logger.debug(f"To target week: {target_week_start_utc} to {target_week_end_utc}")

        # Get previous week's schedules
        previous_schedules = Schedule.query.filter(
            Schedule.start_time >= previous_week_start_utc,
            Schedule.start_time < previous_week_end_utc
        ).all()

        if not previous_schedules:
            flash('No schedules found in previous week to copy.')
            return redirect(url_for('calendar'))

        try:
            # Delete existing schedules in target week
            Schedule.query.filter(
                Schedule.start_time >= target_week_start_utc,
                Schedule.start_time < target_week_end_utc
            ).delete()

            # Copy schedules to new week
            time_difference = target_week_start_utc - previous_week_start_utc

            for schedule in previous_schedules:
                new_schedule = Schedule(
                    technician_id=schedule.technician_id,
                    start_time=schedule.start_time + time_difference,
                    end_time=schedule.end_time + time_difference,
                    description=schedule.description,
                    time_off=schedule.time_off,
                    location_id=schedule.location_id
                )
                db.session.add(new_schedule)

            db.session.commit()
            flash(f'Successfully copied {len(previous_schedules)} schedules from previous week!')

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Database error in copy_previous_week_schedules: {str(e)}")
            flash('Error copying schedules from previous week.')

    except Exception as e:
        app.logger.error(f"Error in copy_previous_week_schedules: {str(e)}")
        flash('Error copying schedules from previous week.')

    return redirect(url_for('calendar', week_start=target_week_start_str))

@app.route('/update_timezone', methods=['POST'])
@login_required
def update_timezone():
    timezone = request.form.get('timezone')
    if timezone in pytz.all_timezones:
        current_user.timezone = timezone
        db.session.commit()
        flash('Timezone updated successfully!')
    else:
        flash('Invalid timezone')
    return redirect(request.referrer or url_for('calendar'))

@app.route('/admin/dashboard', methods=['GET', 'POST'])
@login_required
def admin_dashboard():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    users = User.query.all()
    quick_links = QuickLink.query.order_by(QuickLink.order.asc()).all()
    form = AdminUserForm()
    edit_form = EditUserForm()
    return render_template('admin/dashboard.html', 
                         users=users, 
                         form=form, 
                         edit_form=edit_form,
                         quick_links=quick_links)

@app.route('/admin/create_user', methods=['POST'])
@login_required
def admin_create_user():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    form = AdminUserForm()
    if form.validate_on_submit():
        try:
            # Convert email to lowercase for case-insensitive searching
            email = form.email.data.lower() if form.email.data else ""
            username = form.username.data
            
            # Check if user already exists (case-insensitive using PostgreSQL LOWER)
            existing_email_user = User.query.filter(db.func.lower(User.email) == db.func.lower(email)).first()
            if existing_email_user:
                flash('Email already registered.')
                return redirect(url_for('admin_dashboard'))
                
            # Check if username already exists (case-insensitive using PostgreSQL LOWER)
            if username:
                existing_user = User.query.filter(db.func.lower(User.username) == db.func.lower(username)).first()
                if existing_user:
                    flash('Username already registered. Please choose another username.')
                    return redirect(url_for('admin_dashboard'))

            # Create new user
            user = User(
                username=form.username.data,
                email=email,  # Store email in lowercase
                color=form.color.data or '#3498db',  # Default color if none provided
                is_admin=form.is_admin.data,
                timezone=form.timezone.data or 'America/Los_Angeles'  # Default timezone
            )
            user.set_password(form.password.data)
            
            # Handle profile picture upload
            if form.profile_picture.data:
                try:
                    file = form.profile_picture.data
                    if file and file.filename:
                        # Create upload directory if it doesn't exist
                        upload_dir = os.path.join(app.static_folder, 'uploads', 'profile_pictures')
                        if not os.path.exists(upload_dir):
                            os.makedirs(upload_dir, exist_ok=True)
                        
                        # Generate timestamp-based filename
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        filename = f"{timestamp}_{secure_filename(user.username)}.{file.filename.rsplit('.', 1)[1].lower()}"
                        filepath = os.path.join(upload_dir, filename)
                        
                        # Save the file
                        file.save(filepath)
                        
                        # Store relative path in database
                        user.profile_picture = f"uploads/profile_pictures/{filename}"
                        app.logger.info(f"Profile picture saved for user {user.username}: {user.profile_picture}")
                        
                except Exception as e:
                    app.logger.error(f"Error uploading profile picture: {str(e)}")
                    # Continue user creation even if profile picture upload fails
                    flash('User created successfully, but profile picture upload failed.')

            # Log the creation attempt
            app.logger.info(f"Creating new user with username: {user.username}, email: {user.email}, timezone: {user.timezone}")

            db.session.add(user)
            db.session.commit()
            flash('User created successfully!')

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error creating user: {str(e)}")
            flash('Error creating user. Please check the form and try again.')
    else:
        # Log form validation errors
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{field}: {error}")
                app.logger.error(f"Form validation error - {field}: {error}")

    return redirect(url_for('admin_dashboard'))

@app.route('/admin/edit_user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def admin_edit_user(user_id):
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    user = User.query.get_or_404(user_id)
    form = EditUserForm()

    # Debug log for incoming request
    app.logger.debug(f"Edit user request for user_id {user_id}")
    app.logger.debug(f"Request method: {request.method}")
    app.logger.debug(f"Form data: {request.form}")

    if request.method == 'GET':
        form.username.data = user.username
        form.email.data = user.email
        form.color.data = user.color
        form.is_admin.data = user.is_admin
        form.timezone.data = user.timezone
        return render_template('admin/edit_user.html', 
                            user=user,
                            form=form)

    if request.method == 'POST':
        # Get form data directly from request.form
        username = request.form.get('username')
        email = request.form.get('email')
        
        # Convert email to lowercase for case-insensitivity
        if email:
            email = email.lower()
            
        color = request.form.get('color')
        password = request.form.get('password')
        timezone = request.form.get('timezone')
        is_admin = 'is_admin' in request.form

        # Handle profile picture upload
        profile_picture_path = None
        app.logger.debug(f"Form files: {list(request.files.keys())}")
        
        if 'profile_picture' in request.files:
            file = request.files['profile_picture']
            app.logger.debug(f"Profile picture file: {file}, filename: {file.filename}")
            
            if file and file.filename != '':
                app.logger.debug(f"Processing file upload: {file.filename}")
                # Check file extension
                allowed_extensions = {'jpg', 'jpeg', 'png', 'webp'}
                file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                app.logger.debug(f"File extension: {file_ext}")
                
                if '.' in file.filename and file_ext in allowed_extensions:
                    filename = secure_filename(file.filename)
                    # Add timestamp to avoid name conflicts
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
                    filename = f"{timestamp}{filename}"
                    
                    # Ensure uploads directory exists
                    upload_folder = os.path.join(app.static_folder, 'uploads', 'profile_pictures')
                    os.makedirs(upload_folder, exist_ok=True)
                    app.logger.debug(f"Upload folder: {upload_folder}")
                    
                    # Save the file
                    file_path = os.path.join(upload_folder, filename)
                    try:
                        file.save(file_path)
                        # Store relative path for database
                        profile_picture_path = f"uploads/profile_pictures/{filename}"
                        app.logger.debug(f"Profile picture saved successfully: {profile_picture_path}")
                        flash('Profile picture uploaded successfully!')
                    except Exception as save_error:
                        app.logger.error(f"Error saving profile picture: {save_error}")
                        flash('Error uploading profile picture. Please try again.')
                        return redirect(url_for('admin_edit_user', user_id=user_id))
                else:
                    app.logger.warning(f"Invalid file type: {file_ext}")
                    flash('Invalid file type. Please use JPG, PNG, or WEBP files.')
                    return redirect(url_for('admin_edit_user', user_id=user_id))
            else:
                app.logger.debug("No file selected or empty filename")
        else:
            app.logger.debug("No profile_picture field in request.files")

        app.logger.debug(f"Processed form data: username={username}, email={email}, color={color}, timezone={timezone}, is_admin={is_admin}, profile_picture={profile_picture_path}")

        try:
            # Check if username is already taken by another user (case-insensitive using PostgreSQL LOWER)
            if username:
                username_conflict = User.query.filter(
                    User.id != user_id,
                    db.func.lower(User.username) == db.func.lower(username)
                ).first()
                if username_conflict:
                    flash(f'Username "{username}" is already taken. Please use a different username.')
                    return redirect(url_for('admin_dashboard'))
            
            # Check if email is already taken by another user (case-insensitive using PostgreSQL LOWER)
            if email:
                email_conflict = User.query.filter(
                    User.id != user_id,
                    db.func.lower(User.email) == db.func.lower(email)
                ).first()
                if email_conflict:
                    flash(f'Email "{email}" is already registered to another user.')
                    return redirect(url_for('admin_dashboard'))
            
            # Update user fields
            user.username = username
            # Store email in lowercase for case-insensitive handling
            user.email = email.lower() if email else ""
            user.color = color
            user.is_admin = is_admin
            user.timezone = timezone

            # Update profile picture if uploaded
            if profile_picture_path:
                # Remove old profile picture if it exists
                if user.profile_picture:
                    old_file_path = os.path.join(app.static_folder, user.profile_picture)
                    if os.path.exists(old_file_path):
                        try:
                            os.remove(old_file_path)
                            app.logger.debug(f"Removed old profile picture: {old_file_path}")
                        except Exception as e:
                            app.logger.warning(f"Could not remove old profile picture: {e}")
                
                user.profile_picture = profile_picture_path

            if password:
                user.set_password(password)

            # Commit changes and verify
            db.session.commit()

            # Verify the changes were saved
            updated_user = User.query.get(user_id)
            app.logger.debug(f"Updated user values: username={updated_user.username}, email={updated_user.email}, color={updated_user.color}, timezone={updated_user.timezone}, is_admin={updated_user.is_admin}")

            flash('User updated successfully!')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating user: {str(e)}")
            flash('Error updating user. Please try again.')

    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete_user/<int:user_id>')
@login_required
def admin_delete_user(user_id):
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    if current_user.id == user_id:
        flash('Cannot delete your own account.')
        return redirect(url_for('admin_dashboard'))

    user = User.query.get_or_404(user_id)
    
    # Prevent deletion of the System user
    if user.username == "System":
        flash('Cannot delete the System user as it is required for system operations.')
        return redirect(url_for('admin_dashboard'))
    
    try:
        # Get a special system user to reassign content to
        # Create one if it doesn't exist
        system_user = User.query.filter_by(username="System").first()
        if not system_user:
            system_user = User(
                username="System",
                email="system@example.com",
                is_admin=False
            )
            system_user.set_password(''.join(random.choice(string.ascii_letters + string.digits) for _ in range(20)))
            db.session.add(system_user)
            db.session.flush()  # Get the ID without committing
        
        # Reassign ticket comments to system user
        from models import TicketComment
        ticket_comments = TicketComment.query.filter_by(user_id=user_id).all()
        for comment in ticket_comments:
            comment.user_id = system_user.id
        
        # Reassign ticket history entries to system user
        from models import TicketHistory
        ticket_history = TicketHistory.query.filter_by(user_id=user_id).all()
        for history in ticket_history:
            history.user_id = system_user.id
        
        # Reassign tickets created by this user
        from models import Ticket
        created_tickets = Ticket.query.filter_by(created_by=user_id).all()
        for ticket in created_tickets:
            ticket.created_by = system_user.id
        
        # Remove ticket assignments
        assigned_tickets = Ticket.query.filter_by(assigned_to=user_id).all()
        for ticket in assigned_tickets:
            ticket.assigned_to = None
            
        # Delete associated schedules
        Schedule.query.filter_by(technician_id=user_id).delete()
        
        # Now delete the user
        db.session.delete(user)
        db.session.commit()
        flash('User and associated data deleted successfully!')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting user.')
        app.logger.error(f"Error deleting user: {str(e)}")

    return redirect(url_for('admin_dashboard'))

@app.route('/personal_schedule')
@login_required
def personal_schedule():
    week_start = request.args.get('week_start')
    if week_start:
        week_start = datetime.strptime(week_start, '%Y-%m-%d')
        week_start = current_user.get_timezone_obj().localize(
            week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        )
    else:
        week_start = datetime.now(current_user.get_timezone_obj())
        week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start -= timedelta(days=week_start.weekday())

    # Convert to UTC for database query
    week_start_utc = week_start.astimezone(pytz.UTC)
    week_end_utc = (week_start + timedelta(days=7)).astimezone(pytz.UTC)

    # Query schedules in UTC
    schedules = Schedule.query.filter(
        Schedule.technician_id == current_user.id,
        Schedule.start_time >= week_start_utc,
        Schedule.start_time < week_end_utc
    ).order_by(Schedule.start_time).all()

    # Convert schedule times to user's timezone
    user_tz = current_user.get_timezone_obj()
    for schedule in schedules:
        # Ensure times are timezone-aware in UTC
        if schedule.start_time.tzinfo is None:
            schedule.start_time = pytz.UTC.localize(schedule.start_time)
        if schedule.end_time.tzinfo is None:
            schedule.end_time = pytz.UTC.localize(schedule.end_time)

        # Special handling for all-day time-off events to prevent timezone date shifting
        if schedule.time_off and schedule.all_day:
            # For all-day events, we need to determine the intended calendar date
            # Since existing entries were created in Chicago time, we reverse-engineer the date
            utc_time = schedule.start_time.astimezone(pytz.UTC)
            
            # Try to determine the original calendar date by checking common US timezones
            chicago_tz = pytz.timezone('America/Chicago')
            pacific_tz = pytz.timezone('America/Los_Angeles')
            
            chicago_display = utc_time.astimezone(chicago_tz)
            pacific_display = utc_time.astimezone(pacific_tz)
            
            # If the UTC time matches Chicago midnight conversion pattern, use Chicago date
            chicago_midnight = chicago_tz.localize(datetime.combine(chicago_display.date(), time(0, 0)))
            if utc_time == chicago_midnight.astimezone(pytz.UTC):
                intended_date = chicago_display.date()
                app.logger.debug(f"Mobile personal schedule all-day OOO {schedule.id}: Detected Chicago-created entry for {intended_date}")
            else:
                # Otherwise, use the current user's timezone date
                intended_date = utc_time.astimezone(user_tz).date()
                app.logger.debug(f"Mobile personal schedule all-day OOO {schedule.id}: Using user timezone date {intended_date}")
            
            # Display as all-day in user's timezone for the intended date
            schedule.start_time = user_tz.localize(datetime.combine(intended_date, time(0, 0)))
            schedule.end_time = user_tz.localize(datetime.combine(intended_date, time(23, 59)))
            app.logger.debug(f"Mobile personal schedule all-day display fix for schedule {schedule.id}: {intended_date} → {schedule.start_time} to {schedule.end_time}")
        else:
            schedule.start_time = schedule.start_time.astimezone(user_tz)
            schedule.end_time = schedule.end_time.astimezone(user_tz)

    form = ScheduleForm()
    form.technician.choices = [(current_user.id, current_user.username)]
    form.technician.data = current_user.id
    
    # Set up location choices
    locations = Location.query.filter_by(active=True).order_by(Location.name).all()
    form.location_id.choices = [(l.id, l.name) for l in locations]
    # Add an empty choice if no locations exist
    if not locations:
        form.location_id.choices = [(0, 'No locations available')]

    # Debug mobile detection
    print(f"is_mobile_device() in personal_schedule: {is_mobile_device()}")
    is_mobile = is_mobile_device()  # Force evaluation
    print(f"is_mobile value in personal_schedule: {is_mobile}")
    
    # Check if user is on a mobile device
    if is_mobile_device():
        print("Using mobile template for personal schedule")
        return render_template('mobile_personal_schedule.html', 
                            schedules=schedules,
                            week_start=week_start,
                            week_end=week_start + timedelta(days=7),
                            form=form,
                            locations=locations,  # Add locations to mobile template
                            today=datetime.now(current_user.get_timezone_obj()),
                            user_timezone=str(current_user.get_timezone_obj()),
                            datetime=datetime,
                            timedelta=timedelta,
                            personal_view=True)
    else:
        return render_template('personal_schedule.html', 
                            schedules=schedules,
                            week_start=week_start,
                            week_end=week_start + timedelta(days=7),
                            form=form,
                            today=datetime.now(current_user.get_timezone_obj()),
                            user_timezone=str(current_user.get_timezone_obj()),
                            datetime=datetime,
                            timedelta=timedelta,
                            personal_view=True)

@app.route('/admin/export_schedules')
@login_required
def export_schedules():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        if not start_date or not end_date:
            flash('Please select both start and end dates.')
            return redirect(url_for('admin_dashboard'))

        # Convert dates to UTC datetime objects using admin user's timezone for date range
        admin_tz = pytz.timezone(current_user.get_timezone())
        start_datetime = admin_tz.localize(
            datetime.strptime(start_date, '%Y-%m-%d')
        ).astimezone(pytz.UTC)
        end_datetime = (admin_tz.localize(
            datetime.strptime(end_date, '%Y-%m-%d')
        ) + timedelta(days=1)).astimezone(pytz.UTC)

        # Create a new Excel workbook
        wb = Workbook()
        # Remove the default sheet
        wb.remove(wb.active)

        # Get all users sorted by username
        users = User.query.order_by(User.username).all()

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        for user in users:
            # Get user's schedules for the date range
            schedules = Schedule.query.filter(
                Schedule.technician_id == user.id,
                Schedule.start_time >= start_datetime,
                Schedule.start_time < end_datetime
            ).order_by(Schedule.start_time).all()

            if not schedules:
                continue

            # Create a new worksheet for each user
            ws = wb.create_sheet(title=user.username[:31])  # Excel limits sheet names to 31 chars

            # Calculate total hours using the user's individual timezone
            user_tz = pytz.timezone(user.get_timezone())
            
            # Write header with timezone information
            ws['A1'] = f'Schedule Export - {user.username}'
            ws['A2'] = f'Period: {start_date} to {end_date}'
            ws['A3'] = f'Timezone: {user.get_timezone()} ({user_tz.zone})'
            total_minutes = 0

            for schedule in schedules:
                start_time = schedule.start_time.astimezone(user_tz)
                end_time = schedule.end_time.astimezone(user_tz)
                duration = (end_time - start_time).total_seconds() / 60
                total_minutes += duration

            total_hours = total_minutes // 60
            ws['A5'] = 'Total Hours:'
            ws['B5'] = f"{total_hours:.0f}:00:00"

            # Write column headers
            headers = ['Day', 'Date', 'Clock In', 'Clock Out', 'Total', 'Type', 'Notes']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill

            row = 8
            date_cursor = start_datetime.astimezone(user_tz)
            end_date = end_datetime.astimezone(user_tz)

            while date_cursor.date() < end_date.date():
                day_schedules = [s for s in schedules 
                               if s.start_time.astimezone(user_tz).date() == date_cursor.date()]

                if day_schedules:
                    for schedule in sorted(day_schedules, key=lambda s: s.start_time):
                        start_time = schedule.start_time.astimezone(user_tz)
                        end_time = schedule.end_time.astimezone(user_tz)
                        minutes = int((end_time - start_time).total_seconds() / 60)
                        hours = minutes // 60

                        # Determine the type and notes
                        entry_type = "Time Off" if schedule.time_off else "Work"
                        notes = []
                        if schedule.location:
                            notes.append(f"Location: {schedule.location.name}")
                        if schedule.time_off:
                            notes.append("TIME OFF")
                        if schedule.description:
                            if "ON-CALL" in schedule.description.upper():
                                notes.append("ON-CALL")
                            elif "PLEX" in schedule.description.upper():
                                notes.append("PLEX")
                            else:
                                notes.append(schedule.description)

                        # Write schedule data
                        ws.cell(row=row, column=1).value = start_time.strftime('%A')
                        ws.cell(row=row, column=2).value = start_time.strftime('%-m/%-d/%Y')
                        ws.cell(row=row, column=3).value = start_time.strftime('%-I:%M %p')
                        ws.cell(row=row, column=4).value = end_time.strftime('%-I:%M %p')
                        ws.cell(row=row, column=5).value = f"{hours}:00"
                        ws.cell(row=row, column=6).value = entry_type
                        ws.cell(row=row,column=7).value = " | ".join(notes) if notes else ""
                        row += 1
                else:
                    # Write empty row for days with no schedule
                    ws.cell(row=row, column=1).value= date_cursor.strftime('%A')
                    ws.cell(row=row, column=2).value = date_cursor.strftime('%-m/%-d/%Y')
                    ws.cell(row=row, column=3).value = "0"
                    ws.cell(row=row, column=4).value = "0"
                    ws.cell(row=row, column=5).value = "0:00"
                    row += 1

                date_cursor += timedelta(days=1)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'timesheets_{start_date}_to_{end_date}.xlsx'
        )

    except Exception as e:
        app.logger.error(f"Error exporting schedules: {str(e)}")
        flash('Error exporting schedules. Please try again.')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/quick_links')
@login_required
def admin_quick_links():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    form = QuickLinkForm()
    quick_links = QuickLink.query.order_by(QuickLink.order.asc()).all()
    return render_template('admin/quick_links.html', quick_links=quick_links, form=form)

@app.route('/admin/quick_links/create', methods=['POST'])
@login_required
def admin_create_quick_link():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    try:
        # Create new quick link
        link = QuickLink(
            title=request.form.get('title'),
            url=request.form.get('url'),
            icon=request.form.get('icon', 'link'),
            category=request.form.get('category'),
            order=request.form.get('order', 0)
        )

        db.session.add(link)
        db.session.commit()
        flash('Quick link created successfully!')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error creating quick link: {str(e)}")
        flash('Error creating quick link. Please try again.')

    return redirect(url_for('admin_quick_links'))

@app.route('/admin/quick_links/edit/<int:link_id>', methods=['POST'])
@login_required
def admin_edit_quick_link(link_id):
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    link = QuickLink.query.get_or_404(link_id)
    try:
        link.title = request.form.get('title')
        link.url = request.form.get('url')
        link.icon = request.form.get('icon')
        link.category = request.form.get('category')
        link.order = request.form.get('order', 0)

        db.session.commit()
        flash('Quick link updated successfully!')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error updating quick link: {str(e)}")
        flash('Error updating quick link. Please try again.')

    return redirect(url_for('admin_quick_links'))

@app.route('/admin/quick_links/delete/<int:link_id>')
@login_required
def admin_delete_quick_link(link_id):
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    link = QuickLink.query.get_or_404(link_id)
    try:
        db.session.delete(link)
        db.session.commit()
        flash('Quick link deleted successfully!')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting quick link: {str(e)}")
        flash('Error deleting quick link.')

    return redirect(url_for('admin_quick_links'))

@app.route('/admin/quick_links/reorder', methods=['POST'])
@login_required
def admin_reorder_quick_links():
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Access denied'}), 403

    try:
        new_order = request.json
        # First, get all links and set a high order number to avoid conflicts
        links = QuickLink.query.all()
        for link in links:
            link.order = 10000 + link.order

        db.session.commit()

        # Then update with new order
        for item in new_order:
            link = QuickLink.query.get(item['id'])
            if link:
                link.order = int(item['order'])

        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error reordering quick links: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

def get_open_tickets(limit=5):
    """Get open tickets for the current user (assigned to or created by)"""
    from models import Ticket, TicketStatus
    from flask_login import current_user
    from app import app
    
    if not current_user.is_authenticated:
        return []
    
    app.logger.debug(f"Fetching up to {limit} active tickets for dashboard display")
    
    # Get all active tickets (open, in progress, pending)
    # This is for the dashboard Active Tickets section
    query = Ticket.query.filter(
        Ticket.status.in_([TicketStatus.OPEN, TicketStatus.IN_PROGRESS, TicketStatus.PENDING])
    ).order_by(
        # Order by priority (highest first) and then creation date (newest first)
        Ticket.priority.desc(),
        Ticket.created_at.desc()
    ).limit(limit)
    
    tickets = query.all()
    app.logger.debug(f"Found {len(tickets)} active tickets for dashboard display")
    
    return tickets

@app.context_processor
def inject_quick_links():
    def get_quick_links():
        return QuickLink.query.order_by(QuickLink.order.asc(), QuickLink.category).all()
    
    def get_user_tickets():
        # Always show active tickets in the dashboard regardless of main content filters
        return get_open_tickets(5)  # Limit to 5 tickets
    
    return dict(
        get_quick_links=get_quick_links,
        get_user_tickets=get_user_tickets
    )

@app.route('/api/upcoming_time_off')
@login_required
def get_upcoming_time_off(for_template=False):
    """
    Get time off entries for the next 2 weeks
    If for_template is True, returns data formatted for template rendering
    Otherwise returns JSON response for API endpoint
    """
    if not current_user.is_authenticated:
        app.logger.debug("Unauthorized access to path: /api/upcoming_time_off")
        return jsonify({'error': 'Authentication required'}), 401
    
    try:
        # Get current time in UTC since our database stores times in UTC
        current_time = datetime.now(pytz.UTC)
        two_weeks_later = current_time + timedelta(days=14)

        # Query for upcoming time off entries
        time_off_entries = (Schedule.query
            .join(User)
            .filter(
                Schedule.start_time >= current_time,
                Schedule.start_time <= two_weeks_later,
                Schedule.time_off == True
            )
            .order_by(Schedule.start_time)
            .all())

        # Format data for template rendering with user color
        if for_template:
            template_entries = []
            for entry in time_off_entries:
                user = User.query.get(entry.technician_id)
                if user:
                    # Convert to user's timezone with special handling for all-day events
                    user_tz = current_user.get_timezone_obj()
                    
                    # Special handling for all-day time-off events to prevent timezone date shifting
                    if entry.time_off and entry.all_day:
                        # For all-day events, determine the intended calendar date
                        utc_time = entry.start_time.astimezone(pytz.UTC)
                        
                        # Try to determine the original calendar date by checking common US timezones
                        chicago_tz = pytz.timezone('America/Chicago')
                        chicago_display = utc_time.astimezone(chicago_tz)
                        
                        # If the UTC time matches Chicago midnight conversion pattern, use Chicago date
                        chicago_midnight = chicago_tz.localize(datetime.combine(chicago_display.date(), time(0, 0)))
                        if utc_time == chicago_midnight.astimezone(pytz.UTC):
                            intended_date = chicago_display.date()
                            app.logger.debug(f"Upcoming time-off OOO {entry.id}: Detected Chicago-created entry for {intended_date}")
                        else:
                            # Otherwise, use the current user's timezone date
                            intended_date = utc_time.astimezone(user_tz).date()
                            app.logger.debug(f"Upcoming time-off OOO {entry.id}: Using user timezone date {intended_date}")
                        
                        # Display as all-day in user's timezone for the intended date
                        start_time = user_tz.localize(datetime.combine(intended_date, time(0, 0)))
                        end_time = user_tz.localize(datetime.combine(intended_date, time(23, 59)))
                        
                        # Format display text for all-day events
                        time_display = f"{start_time.strftime('%b %d')} - OOO All Day"
                        template_entries.append({
                            'username': user.username,
                            'color': user.color,
                            'start_time': time_display,
                            'end_time': '',  # No end time display for all-day events
                            'description': entry.description or 'Out of Office',
                            'profile_picture': user.profile_picture if user.profile_picture else None
                        })
                    else:
                        start_time = entry.start_time.astimezone(user_tz)
                        end_time = entry.end_time.astimezone(user_tz)
                        
                        template_entries.append({
                            'username': user.username,
                            'color': user.color,
                            'start_time': start_time.strftime('%b %d, %I:%M %p'),
                            'end_time': end_time.strftime('%b %d, %I:%M %p'),
                            'description': entry.description,
                            'profile_picture': user.profile_picture if user.profile_picture else None
                        })
            return template_entries
        
        # Original API response formatting
        time_off_data = []
        user_tz = pytz.timezone('America/Los_Angeles')  # Default timezone
        
        # Group entries by username and consolidate consecutive dates
        user_entries = {}
        formatted_entries = []

        for entry in time_off_entries:
            user = User.query.get(entry.technician_id)
            if not user:
                continue
                
            username = user.username
            
            if username not in user_entries:
                user_entries[username] = []

            start_local = entry.start_time.astimezone(user_tz)
            end_local = entry.end_time.astimezone(user_tz)

            user_entries[username].append({
                'start_date': start_local.date(),
                'end_date': end_local.date(),
                'description': entry.description,
                'color': user.color
            })

        # Consolidate consecutive dates for each user
        for username, entries in user_entries.items():
            if not entries:
                continue
                
            entries.sort(key=lambda x: x['start_date'])
            consolidated = []
            current_entry = entries[0]

            for entry in entries[1:]:
                if (entry['start_date'] - current_entry['end_date']).days <= 1:
                    # Consecutive days, extend the current entry
                    current_entry['end_date'] = max(current_entry['end_date'], entry['end_date'])
                else:
                    # Non-consecutive, add current entry and start a new one
                    consolidated.append(current_entry)
                    current_entry = entry

            consolidated.append(current_entry)

            # Format consolidated entries
            for entry in consolidated:
                duration = (entry['end_date'] - entry['start_date']).days + 1
                # Get the user object to access profile picture
                user = User.query.filter_by(username=username).first()
                formatted_entries.append({
                    'username': username,
                    'start_date': entry['start_date'].strftime('%b %d'),
                    'end_date': entry['end_date'].strftime('%b %d'),
                    'duration': f"{duration} day{'s' if duration != 1 else ''}",
                    'description': entry.get('description') or 'Time Off',
                    'color': entry.get('color', '#3498db'),
                    'profile_picture': user.profile_picture if user and user.profile_picture else None
                })

        return jsonify(formatted_entries)
    except Exception as e:
        app.logger.error(f"Error in get_upcoming_time_off: {str(e)}")
        if for_template:
            return []
        return jsonify([])

@app.route('/admin/backup')
@login_required
def admin_backup():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))
    return render_template('admin/backup.html')

@app.route('/admin/backup/download')
@login_required
def download_backup():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    try:
        app.logger.info("Starting backup data collection...")
        
        # Collect data efficiently with progress logging
        backup_data = {}
        
        # Users
        users = User.query.all()
        backup_data['users'] = [user.to_dict() for user in users]
        app.logger.info(f"Collected {len(users)} users")
        
        # Locations  
        locations = Location.query.all()
        backup_data['locations'] = [location.to_dict() for location in locations]
        app.logger.info(f"Collected {len(locations)} locations")
        
        # Schedules (limit to recent ones to prevent freezing)
        schedules = Schedule.query.order_by(Schedule.created_at.desc()).limit(2000).all()
        backup_data['schedules'] = [schedule.to_dict() for schedule in schedules]
        app.logger.info(f"Collected {len(schedules)} schedules (limited to most recent 2000)")
        
        # Quick Links
        quick_links = QuickLink.query.all()
        backup_data['quick_links'] = [link.to_dict() for link in quick_links]
        app.logger.info(f"Collected {len(quick_links)} quick links")
        
        # Ticket Categories
        categories = TicketCategory.query.all()
        backup_data['ticket_categories'] = [category.to_dict() for category in categories]
        app.logger.info(f"Collected {len(categories)} ticket categories")
        
        # Tickets (limit to recent ones to prevent freezing)
        tickets = Ticket.query.order_by(Ticket.created_at.desc()).limit(1000).all()
        backup_data['tickets'] = [ticket.to_dict() for ticket in tickets]
        app.logger.info(f"Collected {len(tickets)} tickets (limited to most recent 1000)")
        
        # Email Settings
        email_settings = EmailSettings.query.all()
        backup_data['email_settings'] = [settings.to_dict() for settings in email_settings]
        app.logger.info(f"Collected {len(email_settings)} email settings")
        
        # Recurring Schedule Templates
        templates = RecurringScheduleTemplate.query.all()
        backup_data['recurring_schedule_templates'] = [template.to_dict() for template in templates]
        app.logger.info(f"Collected {len(templates)} recurring templates")

        # Create the backup file with optimized processing
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        app.logger.info("Starting optimized JSON processing...")
        try:
            # Process in smaller chunks to prevent browser freezing
            def generate_backup_json():
                yield '{'
                
                # Process each section separately with yield points
                sections = list(backup_data.keys())
                for i, section_name in enumerate(sections):
                    yield f'"{section_name}":'
                    
                    # Convert section data in smaller chunks
                    section_data = backup_data[section_name]
                    if isinstance(section_data, list) and len(section_data) > 100:
                        # Process large lists in chunks
                        yield '['
                        for j, item in enumerate(section_data):
                            if j > 0:
                                yield ','
                            yield json.dumps(item, default=str, separators=(',', ':'))
                            # Yield control periodically to prevent freezing
                            if j % 50 == 0 and j > 0:
                                app.logger.debug(f"Processed {j}/{len(section_data)} items in {section_name}")
                        yield ']'
                    else:
                        # Small sections can be processed normally
                        yield json.dumps(section_data, default=str, separators=(',', ':'))
                    
                    # Add comma between sections (except last one)
                    if i < len(sections) - 1:
                        yield ','
                
                yield '}'
                app.logger.info("Backup JSON generation completed")
            
            # Create streaming response to prevent memory issues
            response = Response(
                generate_backup_json(),
                mimetype='application/json',
                headers={
                    'Content-Disposition': f'attachment; filename=backup_{timestamp}.json',
                    'Cache-Control': 'no-cache'
                }
            )

            app.logger.info(f"Backup created successfully - {len(backup_data['schedules'])} schedules, {len(backup_data['tickets'])} tickets")
            return response
            
        except Exception as json_error:
            app.logger.error(f"JSON processing failed: {str(json_error)}")
            raise

    except Exception as e:
        app.logger.error(f"Error creating backup: {str(e)}")
        flash('Error creating backup')
        return redirect(url_for('admin_backup'))

@app.route('/admin/restore', methods=['POST'])
@login_required
def restore_backup():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    if 'backup_file' not in request.files:
        flash('No file uploaded')
        return redirect(url_for('admin_backup'))

    file = request.files['backup_file']
    if not file.filename:
        flash('No file selected')
        return redirect(url_for('admin_backup'))

    try:
        backup_data = json.loads(file.read().decode('utf-8'))
        app.logger.info("Starting backup restoration process...")

        # Initialize mappings for existing data
        existing_users = {user.username.lower(): user for user in User.query.all()}
        existing_locations = {loc.name.lower(): loc for loc in Location.query.all()}
        existing_quick_links = {(link.title.lower(), link.url.lower()): link for link in QuickLink.query.all()}

        try:
            # Process users first
            if 'users' in backup_data:
                app.logger.info("Starting user restoration...")
                for user_data in backup_data['users']:
                    try:
                        username = user_data.get('username')
                        email = user_data.get('email')

                        if not username or not email:
                            app.logger.warning(f"Skipping user with missing data")
                            continue

                        # Check if user exists (case-insensitive)
                        existing_user = existing_users.get(username.lower())
                        if existing_user:
                            app.logger.info(f"User {username} already exists")
                            continue

                        # Create new user
                        user = User(
                            username=username,
                            email=email,
                            password_hash=user_data['password_hash'],
                            color=user_data.get('color', '#3498db'),
                            is_admin=user_data.get('is_admin', False),
                            timezone=user_data.get('timezone', 'America/Los_Angeles')
                        )
                        db.session.add(user)
                        app.logger.info(f"Created new user {username}")

                    except Exception as e:
                        app.logger.error(f"Error processing user {username}: {str(e)}")
                        continue

                db.session.commit()
                app.logger.info("Users committed successfully")

            # Process locations
            if 'locations' in backup_data:
                app.logger.info("Starting location restoration...")
                for loc_data in backup_data['locations']:
                    try:
                        name = loc_data.get('name')
                        if not name:
                            continue

                        # Check if location exists (case-insensitive)
                        existing_location = existing_locations.get(name.lower())
                        if existing_location:
                            app.logger.info(f"Location {name} already exists")
                            continue

                        # Create new location
                        location = Location(
                            name=name,
                            description=loc_data.get('description', ''),
                            active=loc_data.get('active', True)
                        )
                        db.session.add(location)
                        app.logger.info(f"Created new location {name}")

                    except Exception as e:
                        app.logger.error(f"Error processing location {name}: {str(e)}")
                        continue

                db.session.commit()
                app.logger.info("Locations committed successfully")

            # Process schedules
            restored_count = 0
            skipped_count = 0

            if 'schedules' in backup_data:
                app.logger.info("Starting schedule restoration...")

                # Refresh user and location data after commits
                users_by_username = {user.username: user for user in User.query.all()}
                locations_by_name = {loc.name: loc for loc in Location.query.all()}

                for schedule_data in backup_data['schedules']:
                    try:
                        # Get technician by username
                        tech_username = schedule_data.get('technician_username')
                        if not tech_username:
                            app.logger.warning("Schedule missing technician username")
                            skipped_count += 1
                            continue

                        technician = users_by_username.get(tech_username)
                        if not technician:
                            app.logger.warning(f"Could not find technician: {tech_username}")
                            skipped_count += 1
                            continue

                        # Get location by name if specified
                        location_id = None
                        location_name = schedule_data.get('location_name')
                        if location_name:
                            location = locations_by_name.get(location_name)
                            if location:
                                location_id = location.id

                        start_time = datetime.fromisoformat(schedule_data['start_time'])
                        end_time = datetime.fromisoformat(schedule_data['end_time'])

                        # Check for existing schedule with same technician, time, and location
                        existing_schedule = Schedule.query.filter_by(
                            technician_id=technician.id,
                            start_time=start_time,
                            end_time=end_time,
                            location_id=location_id
                        ).first()

                        if existing_schedule:
                            app.logger.info(f"Schedule already exists for {tech_username} at {start_time}")
                            skipped_count += 1
                            continue

                        schedule = Schedule(
                            technician_id=technician.id,
                            location_id=location_id,
                            start_time=start_time,
                            end_time=end_time,
                            description=schedule_data.get('description'),
                            time_off=schedule_data.get('time_off', False)
                        )
                        db.session.add(schedule)
                        restored_count += 1

                    except Exception as e:
                        app.logger.error(f"Error processing schedule: {str(e)}")
                        skipped_count += 1
                        continue

                db.session.commit()
                app.logger.info(f"Schedules committed successfully. Restored: {restored_count}, Skipped: {skipped_count}")

            # Process quick links if present
            if 'quick_links' in backup_data:
                app.logger.info("Starting quick links restoration...")
                for link_data in backup_data['quick_links']:
                    try:
                        title = link_data.get('title', '')
                        url = link_data.get('url', '')
                        
                        # Skip if title or url is missing
                        if not title or not url:
                            app.logger.warning(f"Skipping quick link with missing title or URL")
                            continue

                        # Check if link already exists (case-insensitive)
                        title_lower = title.lower()
                        url_lower = url.lower()
                        
                        if (title_lower, url_lower) in existing_quick_links:
                            app.logger.info(f"Quick link already exists: {title}")
                            continue

                        # Create new quick link
                        link = QuickLink(
                            title=title,
                            url=url,
                            icon=link_data.get('icon', 'link'),
                            category=link_data.get('category', 'Uncategorized'),
                            order=link_data.get('order', 0)
                        )
                        db.session.add(link)
                        
                        # Add to existing_quick_links to prevent duplicates within this restore
                        existing_quick_links[(title_lower, url_lower)] = link
                        app.logger.info(f"Created new quick link: {title}")

                    except Exception as e:
                        app.logger.error(f"Error processing quick link {link_data.get('title', 'Unknown')}: {str(e)}")
                        continue

                db.session.commit()
                app.logger.info("Quick links committed successfully")
                
            # Process ticket categories if present
            if 'ticket_categories' in backup_data:
                app.logger.info("Starting ticket categories restoration...")
                existing_categories = {cat.name.lower(): cat for cat in TicketCategory.query.all()}
                
                for category_data in backup_data['ticket_categories']:
                    try:
                        name = category_data.get('name')
                        if not name:
                            continue
                            
                        # Check if category already exists
                        if name.lower() in existing_categories:
                            app.logger.info(f"Ticket category {name} already exists")
                            continue
                            
                        # Create new category
                        category = TicketCategory(
                            name=name,
                            description=category_data.get('description', ''),
                            icon=category_data.get('icon', 'help-circle'),
                            priority_level=category_data.get('priority_level', 0)
                        )
                        db.session.add(category)
                        app.logger.info(f"Created new ticket category: {name}")
                        
                    except Exception as e:
                        app.logger.error(f"Error processing ticket category: {str(e)}")
                        continue
                        
                db.session.commit()
                app.logger.info("Ticket categories committed successfully")
                
            # Process tickets if present
            tickets_restored = 0
            tickets_skipped = 0
            
            if 'tickets' in backup_data:
                app.logger.info("Starting tickets restoration...")
                
                # Refresh reference data
                users_by_username = {user.username: user for user in User.query.all()}
                categories_by_name = {cat.name: cat for cat in TicketCategory.query.all()}
                
                for ticket_data in backup_data['tickets']:
                    try:
                        title = ticket_data.get('title')
                        description = ticket_data.get('description')
                        
                        if not title or not description:
                            app.logger.warning("Ticket missing title or description")
                            tickets_skipped += 1
                            continue
                            
                        # Find category by name
                        category_name = ticket_data.get('category_name')
                        if not category_name or category_name not in categories_by_name:
                            app.logger.warning(f"Could not find category: {category_name}")
                            tickets_skipped += 1
                            continue
                            
                        # Find creator by username
                        creator_username = ticket_data.get('creator_username')
                        if not creator_username or creator_username not in users_by_username:
                            app.logger.warning(f"Could not find creator: {creator_username}")
                            tickets_skipped += 1
                            continue
                            
                        # Find assigned user by username if present
                        assigned_to = None
                        assigned_username = ticket_data.get('assigned_username')
                        if assigned_username and assigned_username in users_by_username:
                            assigned_to = users_by_username[assigned_username].id
                            
                        # Parse dates
                        created_at = None
                        if ticket_data.get('created_at'):
                            created_at = datetime.fromisoformat(ticket_data['created_at'])
                            
                        updated_at = None
                        if ticket_data.get('updated_at'):
                            updated_at = datetime.fromisoformat(ticket_data['updated_at'])
                            
                        due_date = None
                        if ticket_data.get('due_date'):
                            due_date = datetime.fromisoformat(ticket_data['due_date'])
                            
                        # Check if ticket with the original ID already exists
                        ticket_id = ticket_data.get('id')
                        existing_ticket = None
                        if ticket_id:
                            existing_ticket = Ticket.query.get(ticket_id)
                            
                        if existing_ticket:
                            app.logger.info(f"Ticket ID {ticket_id} already exists, skipping")
                            tickets_skipped += 1
                            continue
                            
                        # Create the ticket with its original ID to properly restore it
                        from sqlalchemy import text
                        
                        # First create the ticket without ID
                        ticket = Ticket(
                            title=title,
                            description=description,
                            category_id=categories_by_name[category_name].id,
                            status=ticket_data.get('status', TicketStatus.OPEN),
                            priority=ticket_data.get('priority', 0),
                            assigned_to=assigned_to,
                            created_by=users_by_username[creator_username].id,
                            created_at=created_at,
                            updated_at=updated_at,
                            due_date=due_date,
                            archived=ticket_data.get('archived', False)
                        )
                        
                        # Set the ID to match the original if provided
                        if ticket_id:
                            # Use raw SQL to set the ID explicitly
                            db.session.execute(text("ALTER SEQUENCE ticket_id_seq RESTART WITH :next_id"), 
                                            {"next_id": int(ticket_id) + 1})
                            ticket.id = ticket_id
                            
                        db.session.add(ticket)
                        db.session.flush()  # Get the ticket ID before adding comments
                        
                        # Process comments if present
                        if 'comments' in ticket_data:
                            for comment_data in ticket_data['comments']:
                                user_username = comment_data.get('username')
                                if not user_username or user_username not in users_by_username:
                                    # Try to use system user for comments if the original user isn't found
                                    system_user = User.query.filter_by(username="System").first()
                                    if not system_user:
                                        # Create system user if needed
                                        system_user = User(
                                            username="System", 
                                            email="system@example.com",
                                            is_admin=False
                                        )
                                        system_user.set_password(''.join(random.choice(string.ascii_letters + string.digits) for _ in range(20)))
                                        db.session.add(system_user)
                                        db.session.flush()
                                    user_id = system_user.id
                                else:
                                    user_id = users_by_username[user_username].id
                                    
                                # Get original comment ID if available
                                comment_id = comment_data.get('id')
                                
                                # Create the comment
                                comment = TicketComment(
                                    ticket_id=ticket.id,
                                    user_id=user_id,
                                    content=comment_data.get('content', ''),
                                    created_at=datetime.fromisoformat(comment_data['created_at']) if comment_data.get('created_at') else None,
                                    updated_at=datetime.fromisoformat(comment_data['updated_at']) if comment_data.get('updated_at') else None
                                )
                                
                                # Restore the original ID if possible
                                if comment_id:
                                    db.session.execute(text("ALTER SEQUENCE ticket_comment_id_seq RESTART WITH :next_id"), 
                                                    {"next_id": int(comment_id) + 1})
                                    comment.id = comment_id
                                    
                                db.session.add(comment)
                                
                        # Process history entries if present
                        if 'history' in ticket_data:
                            for history_data in ticket_data['history']:
                                user_username = history_data.get('username')
                                if not user_username or user_username not in users_by_username:
                                    # Try to use system user for history entries if the original user isn't found
                                    system_user = User.query.filter_by(username="System").first()
                                    if not system_user:
                                        # Create system user if needed
                                        system_user = User(
                                            username="System", 
                                            email="system@example.com",
                                            is_admin=False
                                        )
                                        system_user.set_password(''.join(random.choice(string.ascii_letters + string.digits) for _ in range(20)))
                                        db.session.add(system_user)
                                        db.session.flush()
                                    user_id = system_user.id
                                else:
                                    user_id = users_by_username[user_username].id
                                
                                # Get original history ID if available
                                history_id = history_data.get('id')
                                
                                # Create the history entry
                                history = TicketHistory(
                                    ticket_id=ticket.id,
                                    user_id=user_id,
                                    action=history_data.get('action', ''),
                                    details=history_data.get('details', ''),
                                    created_at=datetime.fromisoformat(history_data['created_at']) if history_data.get('created_at') else None
                                )
                                
                                # Restore the original ID if possible
                                if history_id:
                                    db.session.execute(text("ALTER SEQUENCE ticket_history_id_seq RESTART WITH :next_id"), 
                                                    {"next_id": int(history_id) + 1})
                                    history.id = history_id
                                    
                                db.session.add(history)
                                
                        tickets_restored += 1
                        
                    except Exception as e:
                        app.logger.error(f"Error processing ticket: {str(e)}")
                        tickets_skipped += 1
                        continue
                        
                db.session.commit()
                app.logger.info(f"Tickets committed successfully. Restored: {tickets_restored}, Skipped: {tickets_skipped}")
                
            # Process email settings if present
            if 'email_settings' in backup_data and backup_data['email_settings']:
                app.logger.info("Restoring email settings...")
                try:
                    settings_data = backup_data['email_settings'][0]
                    settings = EmailSettings.query.first()
                    
                    if not settings:
                        settings = EmailSettings()
                        db.session.add(settings)
                        
                    settings.admin_email_group = settings_data.get('admin_email_group', 'alerts@obedtv.com')
                    settings.notify_on_create = settings_data.get('notify_on_create', True)
                    settings.notify_on_update = settings_data.get('notify_on_update', True)
                    settings.notify_on_delete = settings_data.get('notify_on_delete', True)
                    
                    db.session.commit()
                    app.logger.info("Email settings restored successfully")
                except Exception as e:
                    app.logger.error(f"Error restoring email settings: {str(e)}")
                    
            # Process recurring schedule templates if present
            templates_restored = 0
            templates_skipped = 0
            
            if 'recurring_schedule_templates' in backup_data:
                app.logger.info("Starting recurring schedule templates restoration...")
                
                # Refresh user and location data after commits
                users_by_username = {user.username: user for user in User.query.all()}
                locations_by_name = {loc.name: loc for loc in Location.query.all()}
                
                for template_data in backup_data['recurring_schedule_templates']:
                    try:
                        template_name = template_data.get('template_name')
                        technician_username = template_data.get('technician_username')
                        
                        if not template_name or not technician_username:
                            app.logger.warning("Template missing name or technician")
                            templates_skipped += 1
                            continue
                            
                        # Find technician by username
                        if technician_username not in users_by_username:
                            app.logger.warning(f"Technician {technician_username} not found for template {template_name}")
                            templates_skipped += 1
                            continue
                            
                        technician = users_by_username[technician_username]
                        
                        # Check if template already exists for this technician
                        existing_template = RecurringScheduleTemplate.query.filter_by(
                            technician_id=technician.id,
                            template_name=template_name
                        ).first()
                        
                        if existing_template:
                            app.logger.info(f"Recurring template '{template_name}' for {technician_username} already exists")
                            templates_skipped += 1
                            continue
                            
                        # Find location if specified
                        location = None
                        location_name = template_data.get('location_name')
                        if location_name and location_name in locations_by_name:
                            location = locations_by_name[location_name]
                        
                        # Create new recurring template
                        template = RecurringScheduleTemplate(
                            technician_id=technician.id,
                            template_name=template_name,
                            location_id=location.id if location else None,
                            active=template_data.get('active', True),
                            monday_start=template_data.get('monday_start'),
                            monday_end=template_data.get('monday_end'),
                            tuesday_start=template_data.get('tuesday_start'),
                            tuesday_end=template_data.get('tuesday_end'),
                            wednesday_start=template_data.get('wednesday_start'),
                            wednesday_end=template_data.get('wednesday_end'),
                            thursday_start=template_data.get('thursday_start'),
                            thursday_end=template_data.get('thursday_end'),
                            friday_start=template_data.get('friday_start'),
                            friday_end=template_data.get('friday_end'),
                            saturday_start=template_data.get('saturday_start'),
                            saturday_end=template_data.get('saturday_end'),
                            sunday_start=template_data.get('sunday_start'),
                            sunday_end=template_data.get('sunday_end'),
                            auto_generate=template_data.get('auto_generate', True),
                            weeks_ahead=template_data.get('weeks_ahead', 2)
                        )
                        
                        # Set timestamps if available
                        if 'last_generated' in template_data and template_data['last_generated']:
                            try:
                                template.last_generated = datetime.fromisoformat(template_data['last_generated'].replace('Z', '+00:00'))
                            except:
                                pass
                                
                        db.session.add(template)
                        app.logger.info(f"Created recurring template '{template_name}' for {technician_username}")
                        templates_restored += 1
                        
                    except Exception as e:
                        app.logger.error(f"Error processing recurring template: {str(e)}")
                        templates_skipped += 1
                        continue
                        
                db.session.commit()
                app.logger.info(f"Recurring templates committed successfully. Restored: {templates_restored}, Skipped: {templates_skipped}")

            flash(f'Backup restored successfully! {restored_count} schedules restored, {skipped_count} skipped. {tickets_restored} tickets restored, {tickets_skipped} skipped. {templates_restored} recurring templates restored, {templates_skipped} skipped.')
            app.logger.info("Backup restore completed successfully")

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error in backup restoration: {str(e)}")
            flash('Error restoring backup')
            raise

    except json.JSONDecodeError:
        flash('Invalid backup file format')
    except Exception as e:
        flash('Error restoring backup')
        app.logger.error(f"Unexpected error in restore_backup: {str(e)}")

    return redirect(url_for('admin_backup'))

@app.route('/admin/locations/edit/<int:location_id>', methods=['POST'])
@login_required
def admin_edit_location(location_id):
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    location = Location.query.get_or_404(location_id)
    try:
        # Get form data with proper boolean conversion for active status
        name = request.form.get('name')
        description = request.form.get('description', '')
        # Explicitly check for the checkbox value
        active = 'active' in request.form

        # Validate required fields
        if not name:
            flash('Location name is required.')
            return redirect(url_for('admin_locations'))

        # Track if status changed
        status_changed = location.active != active

        # Update location
        location.name = name
        location.description = description
        location.active = active

        db.session.commit()

        # Log the status change if it occurred
        if status_changed:
            app.logger.info(f"Location '{location.name}' status changed to {'active' if active else 'inactive'}")
            flash(f"Location '{location.name}' has been {'activated' if active else 'deactivated'}.")
        else:
            flash('Location updated successfully!')

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error updating location: {str(e)}")
        flash('Error updating location. Please try again.')

    return redirect(url_for('admin_locations'))

@app.route('/admin/locations/delete/<int:location_id>')
@login_required
def admin_delete_location(location_id):
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    location = Location.query.get_or_404(location_id)

    # Check if location is being used in any schedules
    if location.schedules:
        flash('Cannot delete location that has associated schedules.')
        return redirect(url_for('admin_locations'))

    try:
        db.session.delete(location)
        db.session.commit()
        flash('Location deleted successfully!')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting location: {str(e)}")
        flash('Error deleting location. Please try again.')

    return redirect(url_for('admin_locations'))
@app.route('/admin/email-settings', methods=['GET', 'POST'])
@login_required
def admin_email_settings():
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))

    settings = EmailSettings.query.first()
    if not settings:
        settings = EmailSettings()
        db.session.add(settings)
        db.session.commit()

    form = EmailSettingsForm(obj=settings)

    if form.validate_on_submit():
        try:
            settings.admin_email_group = form.admin_email_group.data
            settings.notify_on_create = form.notify_on_create.data
            settings.notify_on_update = form.notify_on_update.data
            settings.notify_on_delete = form.notify_on_delete.data
            db.session.commit()
            flash('Email settings updated successfully!')
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating email settings: {str(e)}")
            flash('Error updating email settings.')

    return render_template('admin/email_settings.html', form=form)

# ===============================
# RECURRING SCHEDULE ROUTES
# ===============================

@app.route('/admin/recurring-schedules')
@login_required
def recurring_schedules():
    """View all recurring schedule templates"""
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))
    
    templates = RecurringScheduleTemplate.query.all()
    return render_template('admin/recurring_schedules.html', templates=templates)

@app.route('/admin/recurring-schedule/new', methods=['GET', 'POST'])
@login_required
def new_recurring_schedule():
    """Create a new recurring schedule template"""
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))
    
    form = RecurringScheduleForm()
    
    # Populate choices
    form.technician.choices = [(u.id, u.username) for u in User.query.all()]
    form.location_id.choices = [(0, 'No Location')] + [(l.id, l.name) for l in Location.query.filter_by(active=True).all()]
    
    if form.validate_on_submit():
        # Validate time pairs
        validation_errors = form.validate_time_pairs()
        if validation_errors:
            for error in validation_errors:
                flash(error)
            return render_template('admin/recurring_schedule_form.html', form=form, title="New Recurring Schedule")
        
        try:
            template = RecurringScheduleTemplate(
                technician_id=form.technician.data,
                template_name=form.template_name.data,
                location_id=form.location_id.data if form.location_id.data != 0 else None,
                active=form.active.data,
                auto_generate=form.auto_generate.data,
                weeks_ahead=form.weeks_ahead.data,
                
                # Monday
                monday_start=form.monday_start.data if form.monday_start.data else None,
                monday_end=form.monday_end.data if form.monday_end.data else None,
                
                # Tuesday
                tuesday_start=form.tuesday_start.data if form.tuesday_start.data else None,
                tuesday_end=form.tuesday_end.data if form.tuesday_end.data else None,
                
                # Wednesday
                wednesday_start=form.wednesday_start.data if form.wednesday_start.data else None,
                wednesday_end=form.wednesday_end.data if form.wednesday_end.data else None,
                
                # Thursday
                thursday_start=form.thursday_start.data if form.thursday_start.data else None,
                thursday_end=form.thursday_end.data if form.thursday_end.data else None,
                
                # Friday
                friday_start=form.friday_start.data if form.friday_start.data else None,
                friday_end=form.friday_end.data if form.friday_end.data else None,
                
                # Saturday
                saturday_start=form.saturday_start.data if form.saturday_start.data else None,
                saturday_end=form.saturday_end.data if form.saturday_end.data else None,
                
                # Sunday
                sunday_start=form.sunday_start.data if form.sunday_start.data else None,
                sunday_end=form.sunday_end.data if form.sunday_end.data else None,
            )
            
            db.session.add(template)
            db.session.commit()
            
            flash('Recurring schedule template created successfully!')
            return redirect(url_for('recurring_schedules'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error creating recurring schedule template: {str(e)}")
            flash('Error creating recurring schedule template.')
    
    return render_template('admin/recurring_schedule_form.html', form=form, title="New Recurring Schedule")

@app.route('/admin/recurring-schedule/<int:template_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_recurring_schedule(template_id):
    """Edit an existing recurring schedule template"""
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))
    
    template = RecurringScheduleTemplate.query.get_or_404(template_id)
    form = RecurringScheduleForm(obj=template)
    
    # Populate choices
    form.technician.choices = [(u.id, u.username) for u in User.query.all()]
    form.location_id.choices = [(0, 'No Location')] + [(l.id, l.name) for l in Location.query.filter_by(active=True).all()]
    
    # Set current values for GET request
    if request.method == 'GET':
        form.technician.data = template.technician_id
        form.location_id.data = template.location_id or 0
        
        # Pre-populate all time fields
        form.monday_start.data = template.monday_start
        form.monday_end.data = template.monday_end
        form.tuesday_start.data = template.tuesday_start
        form.tuesday_end.data = template.tuesday_end
        form.wednesday_start.data = template.wednesday_start
        form.wednesday_end.data = template.wednesday_end
        form.thursday_start.data = template.thursday_start
        form.thursday_end.data = template.thursday_end
        form.friday_start.data = template.friday_start
        form.friday_end.data = template.friday_end
        form.saturday_start.data = template.saturday_start
        form.saturday_end.data = template.saturday_end
        form.sunday_start.data = template.sunday_start
        form.sunday_end.data = template.sunday_end
    
    if form.validate_on_submit():
        # Validate time pairs
        validation_errors = form.validate_time_pairs()
        if validation_errors:
            for error in validation_errors:
                flash(error)
            return render_template('admin/recurring_schedule_form.html', form=form, title="Edit Recurring Schedule", template=template)
        
        try:
            template.technician_id = form.technician.data
            template.template_name = form.template_name.data
            template.location_id = form.location_id.data if form.location_id.data != 0 else None
            template.active = form.active.data
            template.auto_generate = form.auto_generate.data
            template.weeks_ahead = form.weeks_ahead.data
            
            # Update all day times
            template.monday_start = form.monday_start.data if form.monday_start.data else None
            template.monday_end = form.monday_end.data if form.monday_end.data else None
            template.tuesday_start = form.tuesday_start.data if form.tuesday_start.data else None
            template.tuesday_end = form.tuesday_end.data if form.tuesday_end.data else None
            template.wednesday_start = form.wednesday_start.data if form.wednesday_start.data else None
            template.wednesday_end = form.wednesday_end.data if form.wednesday_end.data else None
            template.thursday_start = form.thursday_start.data if form.thursday_start.data else None
            template.thursday_end = form.thursday_end.data if form.thursday_end.data else None
            template.friday_start = form.friday_start.data if form.friday_start.data else None
            template.friday_end = form.friday_end.data if form.friday_end.data else None
            template.saturday_start = form.saturday_start.data if form.saturday_start.data else None
            template.saturday_end = form.saturday_end.data if form.saturday_end.data else None
            template.sunday_start = form.sunday_start.data if form.sunday_start.data else None
            template.sunday_end = form.sunday_end.data if form.sunday_end.data else None
            
            db.session.commit()
            
            # Update existing auto-generated schedules to match the new template
            updated_count = template.update_existing_schedules()
            if updated_count > 0:
                db.session.commit()
                flash(f'Recurring schedule template updated successfully! {updated_count} existing schedules were updated to match the new template.')
            else:
                flash('Recurring schedule template updated successfully!')
            
            return redirect(url_for('recurring_schedules'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating recurring schedule template: {str(e)}")
            flash('Error updating recurring schedule template.')
    
    return render_template('admin/recurring_schedule_form.html', form=form, title="Edit Recurring Schedule", template=template)

@app.route('/admin/recurring-schedule/<int:template_id>/delete', methods=['POST'])
@login_required
def delete_recurring_schedule(template_id):
    """Delete a recurring schedule template"""
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))
    
    template = RecurringScheduleTemplate.query.get_or_404(template_id)
    
    try:
        db.session.delete(template)
        db.session.commit()
        flash('Recurring schedule template deleted successfully!')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting recurring schedule template: {str(e)}")
        flash('Error deleting recurring schedule template.')
    
    return redirect(url_for('recurring_schedules'))

@app.route('/admin/recurring-schedule/<int:template_id>/generate', methods=['GET', 'POST'])
@login_required
def generate_recurring_schedule(template_id):
    """Generate schedule entries from a recurring template"""
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))
    
    template = RecurringScheduleTemplate.query.get_or_404(template_id)
    
    # Handle GET request by showing confirmation
    if request.method == 'GET':
        return render_template('admin/confirm_generate.html', template=template)
    
    try:
        # Generate schedules with OOO conflict prevention
        schedules = template.generate_schedules()
        
        if not schedules:
            flash('No new schedules to generate. Either all schedules already exist or there are OOO conflicts preventing generation.', 'info')
            return redirect(url_for('recurring_schedules'))
        
        # Count potential OOO conflicts (for information purposes)
        from datetime import date, timedelta
        start_date = date.today()
        end_date = start_date + timedelta(weeks=template.weeks_ahead)
        
        ooo_conflicts = Schedule.query.filter(
            Schedule.technician_id == template.technician_id,
            Schedule.time_off == True,
            Schedule.all_day == True,
            Schedule.start_time >= pytz.UTC.localize(datetime.combine(start_date, datetime.min.time())),
            Schedule.start_time <= pytz.UTC.localize(datetime.combine(end_date, datetime.max.time()))
        ).count()
        
        # Add schedules to database
        for schedule in schedules:
            db.session.add(schedule)
        
        # Update last generated timestamp
        template.last_generated = datetime.now(pytz.UTC)
        
        db.session.commit()
        
        success_msg = f'Generated {len(schedules)} new schedule entries from template "{template.template_name}"!'
        if ooo_conflicts > 0:
            success_msg += f' (Skipped {ooo_conflicts} days due to OOO conflicts)'
        
        flash(success_msg, 'success')
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error generating schedules from template: {str(e)}")
        flash('Error generating schedules from template.', 'error')
    
    return redirect(url_for('recurring_schedules'))

@app.route('/admin/recurring-schedule/<int:template_id>/preview')
@login_required
def preview_recurring_schedule(template_id):
    """Preview what schedules would be generated from a template"""
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))
    
    template = RecurringScheduleTemplate.query.get_or_404(template_id)
    
    try:
        # Generate preview schedules without saving
        schedules = template.generate_schedules()
        
        # Convert to dictionary format for easy display
        schedule_preview = []
        for schedule in schedules:
            # Convert UTC times to user's timezone for display
            user_tz = current_user.get_timezone_obj()
            local_start = schedule.start_time.astimezone(user_tz)
            local_end = schedule.end_time.astimezone(user_tz)
            
            schedule_preview.append({
                'date': local_start.strftime('%Y-%m-%d'),
                'day': local_start.strftime('%A'),
                'start_time': local_start.strftime('%H:%M'),
                'end_time': local_end.strftime('%H:%M'),
                'location': schedule.location.name if schedule.location else 'No Location',
                'description': schedule.description
            })
        
        return render_template('admin/recurring_schedule_preview.html', 
                             template=template, 
                             schedules=schedule_preview)
        
    except Exception as e:
        app.logger.error(f"Error previewing schedules: {str(e)}")
        flash('Error previewing schedules.')
        return redirect(url_for('recurring_schedules'))

@app.route('/admin/recurring-schedule/<int:template_id>/toggle', methods=['POST'])
@login_required
def toggle_recurring_schedule(template_id):
    """Toggle active status of a recurring schedule template"""
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))
    
    template = RecurringScheduleTemplate.query.get_or_404(template_id)
    
    try:
        template.active = not template.active
        db.session.commit()
        
        status = "activated" if template.active else "deactivated"
        flash(f'Recurring schedule template "{template.template_name}" {status}!')
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error toggling recurring schedule template: {str(e)}")
        flash('Error updating recurring schedule template.')
    
    return redirect(url_for('recurring_schedules'))

@app.route('/api/recurring-schedule-auto-generate', methods=['POST'])
@login_required
def auto_generate_recurring_schedules():
    """API endpoint to automatically generate schedules from active templates"""
    if not current_user.is_admin:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        # Find all active templates that need schedule generation
        templates = RecurringScheduleTemplate.query.filter_by(active=True, auto_generate=True).all()
        
        total_generated = 0
        generated_templates = []
        
        for template in templates:
            # Check if it's time to generate new schedules
            # Generate if never generated before, or if last generated was more than a week ago
            should_generate = False
            
            if not template.last_generated:
                should_generate = True
            else:
                # Calculate time since last generation
                time_since_last = datetime.now(pytz.UTC) - template.last_generated
                if time_since_last.days >= 7:  # Generate weekly
                    should_generate = True
            
            if should_generate:
                try:
                    schedules = template.generate_schedules()
                    
                    if schedules:
                        for schedule in schedules:
                            db.session.add(schedule)
                        
                        template.last_generated = datetime.now(pytz.UTC)
                        total_generated += len(schedules)
                        generated_templates.append({
                            'template_name': template.template_name,
                            'technician': template.technician.username,
                            'schedules_generated': len(schedules)
                        })
                        app.logger.info(f"Generated {len(schedules)} schedules for template '{template.template_name}'")
                    else:
                        app.logger.warning(f"No schedules generated for template '{template.template_name}'")
                        
                except Exception as template_error:
                    app.logger.error(f"Error generating schedules for template '{template.template_name}': {str(template_error)}")
                    # Continue with other templates instead of failing completely
                    continue
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'total_generated': total_generated,
            'templates_processed': len(generated_templates),
            'details': generated_templates
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error in auto-generate recurring schedules: {str(e)}")
        return jsonify({'error': 'Error generating schedules'}), 500

@app.route('/admin/recurring-schedules/export')
@login_required
def export_recurring_templates():
    """Export all recurring schedule templates to JSON file"""
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))
    
    try:
        app.logger.info("Starting template export...")
        templates = RecurringScheduleTemplate.query.all()
        app.logger.info(f"Found {len(templates)} templates to export")
        
        # Prepare export data with template details
        export_data = {
            'recurring_schedule_templates': [template.to_dict() for template in templates],
            'export_info': {
                'exported_at': datetime.now(pytz.UTC).isoformat(),
                'exported_by': current_user.username,
                'total_templates': len(templates),
                'format_version': '1.0'
            }
        }
        
        app.logger.info("Serializing template data to JSON...")
        # Create response with compact JSON to reduce size
        response = make_response(json.dumps(export_data, default=str, separators=(',', ':')))
        response.headers['Content-Type'] = 'application/json'
        response.headers['Content-Disposition'] = f'attachment; filename=recurring_templates_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        
        app.logger.info(f"Template export completed successfully - {len(templates)} templates exported by {current_user.username}")
        return response
        
    except Exception as e:
        app.logger.error(f"Error exporting recurring templates: {str(e)}")
        flash('Error exporting templates.')
        return redirect(url_for('recurring_schedules'))

@app.route('/admin/recurring-schedules/import', methods=['POST'])
@login_required
def import_recurring_templates():
    """Import recurring schedule templates from JSON file"""
    if not current_user.is_admin:
        flash('Access denied.')
        return redirect(url_for('calendar'))
    
    if 'template_file' not in request.files:
        flash('No file selected.')
        return redirect(url_for('recurring_schedules'))
    
    file = request.files['template_file']
    if file.filename == '':
        flash('No file selected.')
        return redirect(url_for('recurring_schedules'))
    
    try:
        # Read and parse JSON file
        file_content = file.read().decode('utf-8')
        import_data = json.loads(file_content)
        
        # Extract templates data
        templates_data = import_data.get('recurring_schedule_templates', [])
        if not templates_data:
            flash('No recurring templates found in the file.')
            return redirect(url_for('recurring_schedules'))
        
        # Get current users and locations for reference
        users_by_username = {user.username: user for user in User.query.all()}
        locations_by_name = {loc.name: loc for loc in Location.query.all()}
        
        templates_imported = 0
        templates_skipped = 0
        
        for template_data in templates_data:
            try:
                template_name = template_data.get('template_name')
                technician_username = template_data.get('technician_username')
                technician_name = template_data.get('technician_name')
                technician_id = template_data.get('technician_id')
                
                if not template_name:
                    app.logger.warning("Template missing name")
                    templates_skipped += 1
                    continue
                
                # Find technician by multiple methods
                technician = None
                
                # Method 1: Try by username first
                if technician_username and technician_username in users_by_username:
                    technician = users_by_username[technician_username]
                
                # Method 2: Try by technician_id if username didn't work
                if not technician and technician_id:
                    technician = User.query.get(technician_id)
                
                # Method 3: Try to find by display name (like "Blake G")
                if not technician and technician_name:
                    # Look for users whose username might match the first name
                    first_name = technician_name.split()[0].lower()
                    for user in User.query.all():
                        if user.username.lower().startswith(first_name):
                            technician = user
                            break
                
                if not technician:
                    app.logger.warning(f"Technician not found for template {template_name} (tried: username={technician_username}, name={technician_name}, id={technician_id})")
                    templates_skipped += 1
                    continue
                
                # Check if template already exists for this technician
                existing_template = RecurringScheduleTemplate.query.filter_by(
                    technician_id=technician.id,
                    template_name=template_name
                ).first()
                
                if existing_template:
                    app.logger.info(f"Template '{template_name}' for {technician_username} already exists - skipping")
                    templates_skipped += 1
                    continue
                
                # Find location if specified
                location = None
                location_name = template_data.get('location_name')
                if location_name and location_name in locations_by_name:
                    location = locations_by_name[location_name]
                
                # Extract weekly schedule data
                weekly_schedule = template_data.get('weekly_schedule', {})
                
                # Create new template
                template = RecurringScheduleTemplate(
                    technician_id=technician.id,
                    template_name=template_name,
                    location_id=location.id if location else None,
                    active=template_data.get('active', True),
                    monday_start=weekly_schedule.get('monday', {}).get('start') if weekly_schedule.get('monday', {}).get('working') else None,
                    monday_end=weekly_schedule.get('monday', {}).get('end') if weekly_schedule.get('monday', {}).get('working') else None,
                    tuesday_start=weekly_schedule.get('tuesday', {}).get('start') if weekly_schedule.get('tuesday', {}).get('working') else None,
                    tuesday_end=weekly_schedule.get('tuesday', {}).get('end') if weekly_schedule.get('tuesday', {}).get('working') else None,
                    wednesday_start=weekly_schedule.get('wednesday', {}).get('start') if weekly_schedule.get('wednesday', {}).get('working') else None,
                    wednesday_end=weekly_schedule.get('wednesday', {}).get('end') if weekly_schedule.get('wednesday', {}).get('working') else None,
                    thursday_start=weekly_schedule.get('thursday', {}).get('start') if weekly_schedule.get('thursday', {}).get('working') else None,
                    thursday_end=weekly_schedule.get('thursday', {}).get('end') if weekly_schedule.get('thursday', {}).get('working') else None,
                    friday_start=weekly_schedule.get('friday', {}).get('start') if weekly_schedule.get('friday', {}).get('working') else None,
                    friday_end=weekly_schedule.get('friday', {}).get('end') if weekly_schedule.get('friday', {}).get('working') else None,
                    saturday_start=weekly_schedule.get('saturday', {}).get('start') if weekly_schedule.get('saturday', {}).get('working') else None,
                    saturday_end=weekly_schedule.get('saturday', {}).get('end') if weekly_schedule.get('saturday', {}).get('working') else None,
                    sunday_start=weekly_schedule.get('sunday', {}).get('start') if weekly_schedule.get('sunday', {}).get('working') else None,
                    sunday_end=weekly_schedule.get('sunday', {}).get('end') if weekly_schedule.get('sunday', {}).get('working') else None,
                    auto_generate=template_data.get('auto_generate', True),
                    weeks_ahead=template_data.get('weeks_ahead', 2)
                )
                
                # Set timestamps if available
                if 'last_generated' in template_data and template_data['last_generated']:
                    try:
                        template.last_generated = datetime.fromisoformat(template_data['last_generated'].replace('Z', '+00:00'))
                    except:
                        pass
                
                db.session.add(template)
                app.logger.info(f"Imported template '{template_name}' for {technician.username} ({technician_name or 'N/A'})")
                templates_imported += 1
                
            except Exception as e:
                app.logger.error(f"Error processing template: {str(e)}")
                templates_skipped += 1
                continue
        
        db.session.commit()
        
        flash(f'Import completed! {templates_imported} templates imported, {templates_skipped} skipped.')
        app.logger.info(f"Template import completed by {current_user.username}: {templates_imported} imported, {templates_skipped} skipped")
        
    except json.JSONDecodeError:
        flash('Invalid file format. Please upload a valid JSON file.')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error importing templates: {str(e)}")
        flash('Error importing templates. Please check the file format.')
    
    return redirect(url_for('recurring_schedules'))